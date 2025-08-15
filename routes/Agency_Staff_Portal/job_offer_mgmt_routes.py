# routes/Agency_Staff_Portal/job_offer_mgmt_routes.py
from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, make_response
from flask_login import current_user
from utils.decorators import login_required_with_role, EXECUTIVE_ROLES
from db import get_db_connection
import datetime
import decimal
import mysql.connector
import re
import io
import csv
from collections import OrderedDict 

# Imports for styled Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

JOB_OFFER_REVIEW_ROLES = ['CEO','Founder','Admin']
CLIENT_CONTACT_ROLE_NAME = 'ClientContact'

job_offer_mgmt_bp = Blueprint('job_offer_mgmt_bp', __name__,
                              template_folder='../../../templates',
                              url_prefix='/job-offers')

# --- [UPDATED] - Enhanced Excel Helper Function ---
def _create_styled_excel(report_data, title, header_mapping):
    """Generates a styled Excel report using only OpenPyXL."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    # --- Define Styles ---
    title_font = Font(name='Calibri', size=18, bold=True, color='1F2937')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo color

    # --- Add and Style Title & Subtitle ---
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header_mapping))
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header_mapping))
    subtitle_cell = worksheet.cell(row=2, column=1, value=f"Report generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    subtitle_cell.font = Font(italic=True, color='6B7280')
    subtitle_cell.alignment = center_align
    worksheet.row_dimensions[2].height = 20

    # --- Write and Style Headers ---
    headers = list(header_mapping.keys())
    # Start headers on row 4 to leave a blank row after the subtitle
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    worksheet.row_dimensions[4].height = 25

    # --- Write Data Cell by Cell ---
    data_keys = list(header_mapping.values())
    # Start data on row 5
    for row_num, row_data in enumerate(report_data, 5):
        for col_num, key in enumerate(data_keys, 1):
            cell_value = row_data.get(key, 'N/A')
            # Format dates nicely
            if isinstance(cell_value, (datetime.datetime, datetime.date)):
                cell_value = cell_value.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=col_num, value=cell_value).alignment = left_align

    # --- Adjust Column Widths Dynamically ---
    column_widths = {}
    for col_num, header_title in enumerate(headers, 1):
        column_widths[col_num] = len(str(header_title))
    
    if report_data:
        for row_data in report_data:
            for col_num, key in enumerate(data_keys, 1):
                cell_len = len(str(row_data.get(key, '')))
                if cell_len > column_widths.get(col_num, 0):
                    column_widths[col_num] = cell_len
    
    for col_num, width in column_widths.items():
        # Add a little padding to the width
        worksheet.column_dimensions[get_column_letter(col_num)].width = width + 4
        
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def get_column_options(conn, cursor, table_name, column_name):
    """Dynamically fetches the allowed values for an ENUM or SET column from the database schema."""
    try:
        db_name = conn.database 
        query = """
            SELECT COLUMN_TYPE 
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s AND COLUMN_NAME = %s
        """
        cursor.execute(query, (db_name, table_name, column_name))
        result = cursor.fetchone()
        if result and result.get('COLUMN_TYPE'):
            type_string = result['COLUMN_TYPE']
            values = re.findall(r"'(.*?)'", type_string)
            return values
        return []
    except Exception as e:
        current_app.logger.error(f"Could not fetch options for {table_name}.{column_name}: {e}", exc_info=True)
        return []

def validate_job_offer_data(form_data, is_editing=False, is_client_submission=False):
    """A comprehensive validation helper for the offer form."""
    errors = {}
    
    if is_client_submission:
        if not form_data.get('title', '').strip(): errors['title'] = 'Job title is required.'
        net_salary_str = form_data.get('salary', '').strip()
        if net_salary_str:
            try:
                if decimal.Decimal(net_salary_str) < 0: errors['salary'] = 'Net salary cannot be negative.'
            except decimal.InvalidOperation: errors['salary'] = 'Invalid net salary format.'
    else:
        if not form_data.get('title', '').strip():
            errors['title'] = 'Job title is required.'
        if not is_editing:
            company_selection_mode = form_data.get('company_selection_mode')
            if not form_data.get('company_id'):
                if company_selection_mode == 'existing' and not form_data.get('selected_company_id'):
                    errors['company_id'] = "You must select an existing company."
                elif company_selection_mode == 'new' and not form_data.get('new_company_name', '').strip():
                    errors['new_company_name'] = "New company name cannot be empty."
                elif not company_selection_mode:
                    errors['company_selection_mode'] = "Please select a company option."
        if not form_data.get('category_id'):
            errors['category_id'] = 'Job category is required.'
        net_salary_str = form_data.get('net_salary', '').strip()
        if net_salary_str:
            try:
                if decimal.Decimal(net_salary_str) < 0: errors['net_salary'] = 'Net salary cannot be negative.'
            except decimal.InvalidOperation: errors['net_salary'] = 'Invalid net salary format.'
    
    candidates_needed_str = form_data.get('candidates_needed', '').strip()
    if candidates_needed_str:
        try:
            if int(candidates_needed_str) < 1: errors['candidates_needed'] = 'At least 1 candidate is required.'
        except ValueError: errors['candidates_needed'] = 'Invalid number for candidates needed.'

    max_age_str = form_data.get('max_age', '').strip()
    if max_age_str:
        try:
            if int(max_age_str) < 18: errors['max_age'] = 'Maximum age must be a realistic value (e.g., 18+).'
        except ValueError: errors['max_age'] = 'Invalid number for maximum age.'
        
    return errors


def _create_automated_announcement(db_cursor, source_type, title, content, audience='AllUsers', priority='Normal', posted_by_user_id=None):
    """Helper to create an automated announcement."""
    try:
        sql = "INSERT INTO SystemAnnouncements (Title, Content, Audience, Priority, Source, PostedByUserID, IsActive) VALUES (%s, %s, %s, %s, %s, %s, 1)"
        db_cursor.execute(sql, (title, content, audience, priority, source_type, posted_by_user_id))
        return True
    except Exception as e:
        current_app.logger.error(f"Failed to create automated announcement: {e}", exc_info=True)
        return False

@job_offer_mgmt_bp.route('/submit', methods=['GET', 'POST'])
@login_required_with_role([CLIENT_CONTACT_ROLE_NAME])
def client_submit_job_offer():
    client_company_id = getattr(current_user, 'company_id', None)
    if not client_company_id:
        flash("Your company information could not be found. Please log in again.", "danger")
        return redirect(url_for('login_bp.login'))

    if request.method == 'POST':
        form_data = request.form.to_dict()
        form_data['benefits'] = request.form.getlist('benefits')
        form_data['available_shifts'] = request.form.getlist('available_shifts')
        errors = validate_job_offer_data(form_data, is_client_submission=True)

        if not errors:
            conn = None
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                sql = """INSERT INTO ClientSubmittedJobOffers (
                             CompanyID, SubmittedByUserID, Title, Location, MaxAge, ClosingDate, HasContract,
                             RequiredLanguage, EnglishLevelRequirement, CandidatesNeeded, HiringCadence, WorkLocationType,
                             HiringPlan, ShiftType, AvailableShifts, NetSalary, PaymentTerm, GraduationStatusRequirement,
                             NationalityRequirement, Benefits, IsTransportationProvided, TransportationType, ClientNotes
                         ) VALUES (
                             %(company_id)s, %(user_id)s, %(title)s, %(location)s, %(max_age)s, %(closing_date)s, %(has_contract)s,
                             %(required_language)s, %(english_level)s, %(candidates_needed)s, %(hiring_cadence)s, %(work_location)s,
                             %(hiring_plan)s, %(shift_type)s, %(available_shifts)s, %(salary)s, %(payment_term)s, %(grad_status)s,
                             %(nationality)s, %(benefits)s, %(transport_provided)s, %(transport_type)s, %(client_notes)s
                         )"""
                params = {
                    "company_id": client_company_id,
                    "user_id": current_user.id,
                    "title": form_data.get('title'),
                    "location": form_data.get('location'),
                    "max_age": form_data.get('max_age') or None,
                    "closing_date": form_data.get('closing_date') or None,
                    "has_contract": 1 if form_data.get('has_contract') == 'yes' else 0,
                    "required_language": form_data.get('required_language'),
                    "english_level": form_data.get('english_level'),
                    "candidates_needed": int(form_data.get('candidates_needed', 1)) if form_data.get('candidates_needed') else 1,
                    "hiring_cadence": form_data.get('hiring_cadence'),
                    "work_location": form_data.get('work_location'),
                    "hiring_plan": form_data.get('hiring_plan'),
                    "shift_type": form_data.get('shift_type'),
                    "available_shifts": ",".join(form_data['available_shifts']) if form_data['available_shifts'] else None,
                    "salary": decimal.Decimal(form_data['salary']) if form_data.get('salary') else None,
                    "payment_term": form_data.get('payment_term'),
                    "grad_status": form_data.get('grad_status'),
                    "nationality": form_data.get('nationality'),
                    "benefits": ",".join(form_data['benefits']) if form_data['benefits'] else None,
                    "transport_provided": 1 if form_data.get('transportation') == 'yes' else 0,
                    "transport_type": form_data.get('transport_type') if form_data.get('transportation') == 'yes' else None,
                    "client_notes": form_data.get('client_notes')
                }
                cursor.execute(sql, params)
                conn.commit()
                flash("Job offer submitted for review! Thank you.", "success")
                return redirect(url_for('client_portal_bp.dashboard'))
            except Exception as e:
                if conn: conn.rollback()
                current_app.logger.error(f"Error in client_submit_job_offer: {e}", exc_info=True)
                flash("An error occurred while submitting the job offer. Please contact support.", "danger")
            finally:
                if conn and conn.is_connected(): conn.close()
        else:
            flash("Please correct the errors in the form before submitting.", "warning")
    return redirect(url_for('client_portal_bp.dashboard'))


@job_offer_mgmt_bp.route('/api/offers-for-company/<int:company_id>')
@login_required_with_role(EXECUTIVE_ROLES)
def api_get_offers_for_company(company_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT OfferID, Title FROM JobOffers WHERE CompanyID = %s ORDER BY Title", (company_id,))
    offers = cursor.fetchall()
    cursor.close()
    conn.close()
    return {'offers': offers}

@job_offer_mgmt_bp.route('/api/all-offers')
@login_required_with_role(EXECUTIVE_ROLES)
def api_get_all_offers():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT OfferID, Title FROM JobOffers ORDER BY Title")
    offers = cursor.fetchall()
    cursor.close()
    conn.close()
    return {'offers': offers}

@job_offer_mgmt_bp.route('/dashboard')
@login_required_with_role(JOB_OFFER_REVIEW_ROLES)
def dashboard():
    conn = None
    kpis = {
        'total_open_offers': 0,
        'total_candidates_needed': 0,
        'pending_submissions': 0,
        'new_offers_this_month': 0,
        'avg_time_to_fill': 'N/A',
        'approval_rate': 'N/A',
        'oldest_offer_title': 'No open offers',
        'age_of_oldest_offer': 0,
        'total_applications': 0 # Renamed from 'pending_applications'
    }
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        queries = {
            'total_open_offers': "SELECT COUNT(OfferID) as count FROM JobOffers WHERE Status = 'Open'",
            'total_candidates_needed': "SELECT SUM(CandidatesNeeded) as sum FROM JobOffers WHERE Status = 'Open'",
            'avg_time_to_fill': "SELECT AVG(DATEDIFF(FilledDate, DatePosted)) as avg_days FROM JobOffers WHERE Status = 'Filled' AND FilledDate IS NOT NULL AND DatePosted IS NOT NULL",
            'approval_rate': "SELECT (SUM(CASE WHEN ReviewStatus = 'Approved' THEN 1 ELSE 0 END) / COUNT(SubmissionID)) * 100 as approval_rate FROM ClientSubmittedJobOffers WHERE SubmissionDate >= DATE_SUB(NOW(), INTERVAL 90 DAY) AND ReviewStatus IN ('Approved', 'Rejected')",
            'pending_submissions': "SELECT COUNT(SubmissionID) as count FROM ClientSubmittedJobOffers WHERE ReviewStatus = 'Pending'",
            'new_offers_this_month': "SELECT COUNT(OfferID) as count FROM JobOffers WHERE DatePosted >= DATE_FORMAT(NOW(), '%Y-%m-01')",
            # --- MODIFIED: This now counts ALL applications ---
            'total_applications': "SELECT COUNT(*) as count FROM JobApplications"
        }

        for key, sql in queries.items():
            cursor.execute(sql)
            res = cursor.fetchone()
            if res:
                if key == 'total_candidates_needed':
                    kpis[key] = int(res.get('sum') or 0)
                elif key == 'avg_time_to_fill':
                    kpis[key] = round(res['avg_days'], 1) if res.get('avg_days') is not None else 'N/A'
                elif key == 'approval_rate':
                    kpis[key] = f"{round(res['approval_rate'], 1)}%" if res.get('approval_rate') is not None else 'N/A'
                else:
                    kpis[key] = res.get('count', 0)

        cursor.execute("""
            SELECT jo.Title, c.CompanyName, DATEDIFF(NOW(), jo.DatePosted) as oldest_days 
            FROM JobOffers jo 
            JOIN Companies c ON jo.CompanyID = c.CompanyID 
            WHERE jo.Status = 'Open' 
            ORDER BY jo.DatePosted ASC 
            LIMIT 1
        """)
        oldest = cursor.fetchone()
        if oldest:
            kpis.update({
                'age_of_oldest_offer': oldest['oldest_days'], 
                'oldest_offer_title': oldest['Title'], 
                'oldest_offer_company': oldest['CompanyName']
            })
        else:
             kpis.update({
                'age_of_oldest_offer': 'N/A', 
                'oldest_offer_title': 'No open offers', 
                'oldest_offer_company': ''
             })
        
        return render_template('agency_staff_portal/job_offers/dashboard.html', title="Job Offer Dashboard", kpis=kpis)

    except Exception as e:
        current_app.logger.error(f"Error building KPI dashboard: {e}", exc_info=True)
        flash("Could not load dashboard data.", "danger")
        return render_template('agency_staff_portal/job_offers/dashboard.html', title="Error", kpis=kpis, error=True)
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()

@job_offer_mgmt_bp.route('/')
@login_required_with_role(EXECUTIVE_ROLES)
def list_all_job_offers():
    """
    Lists all job offers, grouped by company, for management.
    """
    offers_by_company = OrderedDict()
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # Fetch all offers, ordered by company name first
        cursor.execute("""
            SELECT 
                jo.OfferID, jo.Title, jo.Status, jo.DatePosted, 
                c.CompanyID, c.CompanyName, 
                jc.CategoryName 
            FROM JobOffers jo 
            JOIN Companies c ON jo.CompanyID = c.CompanyID 
            LEFT JOIN JobCategories jc ON jo.CategoryID = jc.CategoryID 
            ORDER BY c.CompanyName, jo.DatePosted DESC
        """)
        all_offers = cursor.fetchall()

        # Group the flat list into a dictionary of companies
        for offer in all_offers:
            company_id = offer['CompanyID']
            if company_id not in offers_by_company:
                offers_by_company[company_id] = {
                    "company_id": company_id,
                    "company_name": offer['CompanyName'],
                    "offers": []
                }
            offers_by_company[company_id]['offers'].append(offer)

    except Exception as e:
        current_app.logger.error(f"Error fetching and grouping job offers: {e}", exc_info=True)
        flash("Could not load job offers list.", "danger")
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()

    return render_template(
        'agency_staff_portal/job_offers/list_all_job_offers.html', 
        title='Manage Live Offers', 
        offers_by_company=offers_by_company
    )


def get_form_options(conn, cursor):
    """
    Helper to fetch all dynamic options for the job offer form from the JobOffers table schema.
    """
    options = {}
    table = 'JobOffers'
    columns_to_fetch = [
        'PaymentTerm', 'HiringPlan', 'LanguagesType', 'RequiredLevel',
        'GraduationStatusRequirement', 'HiringCadence', 'WorkLocationType', 
        'ShiftType', 'AvailableShifts', 'BenefitsIncluded', 'InterviewType', 
        'Nationality', 'Gender', 'MilitaryStatus', 'Status', 'RequiredLanguages'
    ]
    for col in columns_to_fetch:
        options[col] = get_column_options(conn, cursor, table, col)
    
    return options

@job_offer_mgmt_bp.route('/create-live', methods=['GET', 'POST'])
@login_required_with_role(JOB_OFFER_REVIEW_ROLES)
def staff_direct_create_job_offer():
    errors, companies, categories, form_options = {}, [], [], {}
    form_data = {}

    conn_data = get_db_connection()
    try:
        cursor = conn_data.cursor(dictionary=True)
        cursor.execute("SELECT CompanyID, CompanyName FROM Companies ORDER BY CompanyName")
        companies = cursor.fetchall()
        cursor.execute("SELECT CategoryID, CategoryName FROM JobCategories ORDER BY CategoryName")
        categories = cursor.fetchall()
        form_options = get_form_options(conn_data, cursor)
    except Exception as e:
        current_app.logger.error(f"Error fetching form data for create offer: {e}", exc_info=True)
        flash("Error loading form support data.", "danger")
    finally:
        if conn_data and conn_data.is_connected(): conn_data.close()

    if request.method == 'POST':
        form_data = request.form.to_dict()
        
        # Process multi-select fields into lists for repopulating the form on error
        form_data['graduation_status_requirement_list'] = request.form.getlist('graduation_status_requirement')
        form_data['available_shifts_list'] = request.form.getlist('available_shifts')
        form_data['benefits_included_list'] = request.form.getlist('benefits_included')
        form_data['required_languages_list'] = request.form.getlist('required_languages')

        # Simple validation, can be expanded
        if not form_data.get('title'): errors['title'] = 'Title is required.'
        if not form_data.get('selected_company_id'): errors['company_id'] = 'Company is required.'
        if not form_data.get('category_id'): errors['category_id'] = 'Category is required.'

        if not errors:
            conn_tx = get_db_connection()
            try:
                cursor = conn_tx.cursor(dictionary=True)
                conn_tx.start_transaction()
                
                company_id = form_data.get('selected_company_id')
                cursor.execute("SELECT CompanyName FROM Companies WHERE CompanyID = %s", (company_id,))
                company_result = cursor.fetchone()
                company_name = company_result['CompanyName'] if company_result else 'Unknown Company'

                params = {
                    "company_id": company_id,
                    "posted_by_id": getattr(current_user, 'specific_role_id', None),
                    "category_id": form_data.get('category_id'),
                    "title": form_data.get('title'),
                    "location": form_data.get('location'),
                    "net_salary": decimal.Decimal(form_data['net_salary']) if form_data.get('net_salary') else None,
                    "payment_term": form_data.get('payment_term'),
                    "hiring_plan": form_data.get('hiring_plan'),
                    "max_age": int(form_data['max_age']) if form_data.get('max_age') else None,
                    "has_contract": 1 if form_data.get('has_contract') else 0,
                    "languages_type": form_data.get('languages_type'),
                    "required_languages": ",".join(form_data['required_languages_list']) or None,
                    "required_level": form_data.get('required_level'),
                    "graduation_status_requirement": ",".join(form_data['graduation_status_requirement_list']) or None,
                    "candidates_needed": int(form_data.get('candidates_needed', 1)),
                    "hiring_cadence": form_data.get('hiring_cadence'),
                    "work_location_type": form_data.get('work_location_type'),
                    "shift_type": form_data.get('shift_type'),
                    "available_shifts": ",".join(form_data['available_shifts_list']) or None,
                    "benefits_included": ",".join(form_data['benefits_included_list']) or None,
                    "interview_type": form_data.get('interview_type'),
                    "nationality": form_data.get('nationality'),
                    "gender": form_data.get('gender'),
                    "military_status": form_data.get('military_status'),
                    "closing_date": form_data.get('closing_date') or None,
                    "working_days": form_data.get('working_days'),
                    "working_hours": form_data.get('working_hours'),
                    "experience_requirement": form_data.get('experience_requirement'),
                    "status": form_data.get('status', 'Open')
                }
                
                sql = """
                    INSERT INTO JobOffers (
                        CompanyID, PostedByStaffID, CategoryID, Title, Location, NetSalary, PaymentTerm, HiringPlan,
                        MaxAge, HasContract, LanguagesType, RequiredLanguages, RequiredLevel, GraduationStatusRequirement,
                        CandidatesNeeded, HiringCadence, WorkLocationType, ShiftType, AvailableShifts,
                        BenefitsIncluded, InterviewType, Nationality, Gender, MilitaryStatus, ClosingDate,
                        WorkingDays, WorkingHours, ExperienceRequirement, Status, DatePosted
                    ) VALUES (
                        %(company_id)s, %(posted_by_id)s, %(category_id)s, %(title)s, %(location)s, %(net_salary)s, 
                        %(payment_term)s, %(hiring_plan)s, %(max_age)s, %(has_contract)s,
                        %(languages_type)s, %(required_languages)s, %(required_level)s, %(graduation_status_requirement)s,
                        %(candidates_needed)s, %(hiring_cadence)s, %(work_location_type)s, %(shift_type)s, %(available_shifts)s, 
                        %(benefits_included)s, %(interview_type)s, %(nationality)s, %(gender)s, %(military_status)s,
                        %(closing_date)s, %(working_days)s, %(working_hours)s, %(experience_requirement)s, %(status)s, NOW()
                    )
                """
                cursor.execute(sql, params)
                
                announcement_title = f"New Job: {params['title']} with {company_name}"
                announcement_content = f"A new job offer has been posted directly by staff."
                _create_automated_announcement(
                    cursor, source_type='Automated_Offer', title=announcement_title, content=announcement_content, posted_by_user_id=current_user.id
                )
                
                conn_tx.commit()
                flash(f"New job offer for '{params['title']}' created successfully and announced!", 'success')
                return redirect(url_for('.list_all_job_offers'))

            except Exception as e:
                if conn_tx: conn_tx.rollback()
                current_app.logger.error(f"Error creating live job offer: {e}", exc_info=True)
                flash(f'An error occurred: {e}', 'danger')
        else:
            flash('Please correct the errors shown.', 'warning')
    
    # Default values for GET request
    if not form_data:
        form_data = { 'status': 'Open', 'gender': 'Both', 'military_status': 'Not Applicable' }

    return render_template('agency_staff_portal/job_offers/add_edit_job_offer.html', 
        title='Create Job Offer', form_data=form_data, errors=errors, companies=companies, 
        categories=categories, is_editing_live=False, action_verb="Post Live", form_options=form_options)


@job_offer_mgmt_bp.route('/edit-live/<int:offer_id>', methods=['GET', 'POST'])
@login_required_with_role(JOB_OFFER_REVIEW_ROLES)
def edit_live_job_offer(offer_id):
    errors, companies, categories, form_options = {}, [], [], {}
    form_data_for_template = {}
    
    # --- Step 1: Fetch dependencies (companies, categories, etc.) ---
    conn_deps = get_db_connection()
    try:
        cursor_deps = conn_deps.cursor(dictionary=True)
        # Fetching companies and categories for the form dropdowns
        cursor_deps.execute("SELECT CompanyID, CompanyName FROM Companies ORDER BY CompanyName")
        companies = cursor_deps.fetchall()
        cursor_deps.execute("SELECT CategoryID, CategoryName FROM JobCategories ORDER BY CategoryName")
        categories = cursor_deps.fetchall()
        # Fetching dynamic options (ENUMs, SETs) for the form
        form_options = get_form_options(conn_deps, cursor_deps)
    except Exception as e:
        current_app.logger.error(f"Error fetching dropdown data for edit offer {offer_id}: {e}", exc_info=True)
        flash("Error loading form support data.", "danger")
        if conn_deps and conn_deps.is_connected(): conn_deps.close()
        return redirect(url_for('.list_all_job_offers'))
    finally:
        if conn_deps and conn_deps.is_connected(): conn_deps.close()

    # --- Step 2: Fetch the original offer data to get the OLD status ---
    conn_orig = get_db_connection()
    try:
        cursor_orig = conn_orig.cursor(dictionary=True)
        cursor_orig.execute("SELECT Title, Status FROM JobOffers WHERE OfferID = %s", (offer_id,))
        original_offer = cursor_orig.fetchone()
        if not original_offer:
            flash('Job offer not found.', 'danger')
            return redirect(url_for('.list_all_job_offers'))
        old_status = original_offer['Status']
        original_title = original_offer['Title'] # Store title for announcements
    except Exception as e:
        flash("Could not retrieve original offer data.", "danger")
        current_app.logger.error(f"Error fetching original offer status for {offer_id}: {e}")
        return redirect(url_for('.list_all_job_offers'))
    finally:
        if conn_orig and conn_orig.is_connected(): conn_orig.close()


    if request.method == 'POST':
        # --- Step 3: Process the submitted form ---
        form_data = request.form.to_dict()
        
        # Process multi-select fields
        benefits_list = request.form.getlist('benefits_checkboxes')
        if form_data.get('transportation') == 'yes' and form_data.get('transport_type'):
            benefits_list.append(form_data.get('transport_type'))
        
        form_data['benefits_included'] = benefits_list
        form_data['available_shifts'] = request.form.getlist('available_shifts')
        form_data['required_languages'] = request.form.getlist('required_languages')
        form_data['grad_status_req'] = request.form.getlist('grad_status_req')

        errors = validate_job_offer_data(form_data, is_editing=True)
        
        if not errors:
            conn_update = get_db_connection()
            try:
                cursor_update = conn_update.cursor()
                conn_update.start_transaction()
                
                # Prepare parameters for the UPDATE query
                params = {
                    "category_id": form_data.get('category_id'), "title": form_data.get('title').strip(),
                    "location": form_data.get('location'),
                    "net_salary": decimal.Decimal(form_data['net_salary']) if form_data.get('net_salary') else None,
                    "payment_term": form_data.get('payment_term'), "hiring_plan": form_data.get('hiring_plan'),
                    "max_age": int(form_data['max_age']) if form_data.get('max_age') else None,
                    "has_contract": 1 if form_data.get('has_contract') == 'yes' else 0,
                    "grad_status_req": ",".join(form_data.get('grad_status_req', [])),
                    "languages_type": form_data.get('languages_type'),
                    "required_languages": ",".join(form_data.get('required_languages', [])),
                    "required_level": form_data.get('required_level'),
                    "candidates_needed": int(form_data.get('candidates_needed', 1)),
                    "hiring_cadence": form_data.get('hiring_cadence'),
                    "work_location_type": form_data.get('work_location_type'),
                    "shift_type": form_data.get('shift_type'),
                    "available_shifts": ",".join(form_data.get('available_shifts', [])),
                    "benefits_included": ",".join(form_data.get('benefits_included', [])),
                    "interview_type": form_data.get('interview_type'), "nationality": form_data.get('nationality'),
                    "status": form_data.get('status', 'Open'),
                    "closing_date": form_data.get('closing_date') or None, "offer_id": offer_id,
                    "company_id": form_data.get('CompanyID'),
                    "gender": form_data.get('gender'),
                    "military_status": form_data.get('military_status'),
                    "working_days": form_data.get('working_days'),
                    "working_hours": form_data.get('working_hours'),
                    "experience_requirement": form_data.get('experience_requirement')
                }
                
                # The UPDATE query
                sql = """
                    UPDATE JobOffers SET
                        CompanyID=%(company_id)s, CategoryID=%(category_id)s, Title=%(title)s, Location=%(location)s, NetSalary=%(net_salary)s, PaymentTerm=%(payment_term)s, HiringPlan=%(hiring_plan)s,
                        MaxAge=%(max_age)s, HasContract=%(has_contract)s, GraduationStatusRequirement=%(grad_status_req)s, LanguagesType=%(languages_type)s, 
                        RequiredLanguages=%(required_languages)s, RequiredLevel=%(required_level)s, CandidatesNeeded=%(candidates_needed)s, 
                        HiringCadence=%(hiring_cadence)s, WorkLocationType=%(work_location_type)s, ShiftType=%(shift_type)s, 
                        AvailableShifts=%(available_shifts)s, BenefitsIncluded=%(benefits_included)s, InterviewType=%(interview_type)s, 
                        Nationality=%(nationality)s, Status=%(status)s, ClosingDate=%(closing_date)s, 
                        Gender=%(gender)s, MilitaryStatus=%(military_status)s, WorkingDays=%(working_days)s, WorkingHours=%(working_hours)s, ExperienceRequirement=%(experience_requirement)s,
                        UpdatedAt=NOW()
                    WHERE OfferID = %(offer_id)s
                """
                cursor_update.execute(sql, params)

                # --- Step 4: Compare old and new status and create announcement ---
                new_status = params['status']
                if old_status != new_status:
                    announcement_title = f"Offer Status Updated: {original_title}"
                    announcement_content = f"The status for the job offer '{original_title}' has been changed from '{old_status}' to '{new_status}'."
                    _create_automated_announcement(
                        cursor_update, 
                        source_type='Automated_Status_Change',
                        title=announcement_title, 
                        content=announcement_content,
                        posted_by_user_id=current_user.id
                    )
                    flash(f"Job offer status updated and an announcement was posted.", 'info')

                conn_update.commit()
                flash(f"Job offer '{params['title']}' updated successfully!", 'success')
                return redirect(url_for('.list_all_job_offers'))
                
            except Exception as e:
                if conn_update: conn_update.rollback()
                current_app.logger.error(f"Error updating offer {offer_id}: {e}", exc_info=True)
                flash(f'An error occurred while updating the offer: {e}', 'danger')
            finally:
                if conn_update and conn_update.is_connected(): conn_update.close()
        
        if errors:
            flash('Please correct the form errors.', 'warning')
        
        # If there are errors, we need to repopulate the template with the submitted data
        form_data_for_template = form_data
    
    else: # GET request
        # --- This part populates the form on initial page load ---
        conn_fetch = get_db_connection()
        try:
            cursor_fetch = conn_fetch.cursor(dictionary=True)
            cursor_fetch.execute("SELECT * FROM JobOffers WHERE OfferID = %s", (offer_id,))
            db_data = cursor_fetch.fetchone()
            if not db_data:
                # This check is now redundant due to the check at the top, but safe to keep.
                flash('Job offer not found.', 'danger')
                return redirect(url_for('.list_all_job_offers'))
            
            form_data_for_template = db_data.copy()

            def parse_set_or_text_column(value):
                if isinstance(value, set): return list(value)
                if isinstance(value, bytes): value = value.decode('utf-8')
                if isinstance(value, str): return [v.strip() for v in value.split(',') if v.strip()]
                return []

            # Populate multi-select fields from DB data
            all_benefits_from_db = parse_set_or_text_column(db_data.get('BenefitsIncluded'))
            form_data_for_template['required_languages'] = parse_set_or_text_column(db_data.get('RequiredLanguages'))
            form_data_for_template['available_shifts'] = parse_set_or_text_column(db_data.get('AvailableShifts'))
            form_data_for_template['grad_status_req'] = parse_set_or_text_column(db_data.get('GraduationStatusRequirement'))
            
            # Deconstruct benefits for the form
            dynamic_transport_options = form_options.get('transportation_options', [])
            form_data_for_template['transportation'] = 'no'
            form_data_for_template['transport_type'] = ''
            final_benefits_for_checkboxes = []
            
            for benefit in all_benefits_from_db:
                if benefit in dynamic_transport_options:
                    form_data_for_template['transportation'] = 'yes'
                    form_data_for_template['transport_type'] = benefit
                else:
                    final_benefits_for_checkboxes.append(benefit)
            form_data_for_template['benefits_checkboxes'] = final_benefits_for_checkboxes

            # Format other fields for the form
            form_data_for_template['has_contract'] = 'yes' if db_data.get('HasContract') else 'no'
            form_data_for_template['net_salary'] = str(db_data['NetSalary']) if db_data.get('NetSalary') is not None else ''
            form_data_for_template['closing_date'] = db_data['ClosingDate'].isoformat() if db_data.get('ClosingDate') else ''

        except Exception as e:
            current_app.logger.error(f"Error fetching offer {offer_id} for edit: {e}", exc_info=True)
            flash(f'Error fetching offer details: {e}', 'danger')
            return redirect(url_for('.list_all_job_offers'))
        finally: 
            if conn_fetch and conn_fetch.is_connected():
                if 'cursor_fetch' in locals(): cursor_fetch.close()
                conn_fetch.close()
                
    return render_template('agency_staff_portal/job_offers/add_edit_job_offer.html', 
        title=f'Edit Offer: {form_data_for_template.get("Title", "N/A")}', 
        form_data=form_data_for_template, 
        errors=errors, 
        companies=companies, 
        categories=categories, 
        offer_id=offer_id, 
        is_editing_live=True, 
        action_verb="Update Offer", 
        form_options=form_options)

@job_offer_mgmt_bp.route('/delete-live/<int:offer_id>', methods=['POST'])
@login_required_with_role(JOB_OFFER_REVIEW_ROLES)
def delete_live_job_offer(offer_id):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM JobOffers WHERE OfferID = %s", (offer_id,))
        conn.commit()
        if cursor.rowcount > 0:
            flash('The job offer has been permanently deleted.', 'success')
        else:
            flash('The job offer could not be found or was already deleted.', 'warning')
    except mysql.connector.Error as e:
        if conn: conn.rollback()
        if e.errno == 1451:
            flash('Cannot delete this offer because it has active job applications. Please close the offer instead.', 'danger')
        else:
            current_app.logger.error(f"Database error deleting offer {offer_id}: {e}", exc_info=True)
            flash(f'A database error occurred. Please contact support.', 'danger')
    except Exception as e:
        if conn: conn.rollback()
        current_app.logger.error(f"Generic error deleting offer {offer_id}: {e}", exc_info=True)
        flash(f'An unexpected error occurred.', 'danger')
    finally:
        if conn and conn.is_connected(): 
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()
    return redirect(url_for('.list_all_job_offers'))

@job_offer_mgmt_bp.route('/review-client-submissions')
@login_required_with_role(JOB_OFFER_REVIEW_ROLES)
def list_review_client_submissions():
    conn, submissions, categories_for_dropdown = None, [], []
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT csjo.*, c.CompanyName, u.FirstName as SubmitterFirstName, u.LastName as SubmitterLastName 
            FROM ClientSubmittedJobOffers csjo 
            JOIN Companies c ON csjo.CompanyID = c.CompanyID 
            JOIN Users u ON csjo.SubmittedByUserID = u.UserID
            WHERE csjo.ReviewStatus IN ('Pending', 'NeedsClarification') 
            ORDER BY csjo.SubmissionDate ASC
        """)
        submissions = cursor.fetchall()
        for sub in submissions: 
            benefits_str = sub.get('Benefits', '')
            sub['Benefits_list'] = benefits_str.split(',') if benefits_str else []
            shifts_str = sub.get('AvailableShifts', '')
            sub['AvailableShifts_list'] = shifts_str.split(',') if shifts_str else []
        
        cursor.execute("SELECT CategoryID, CategoryName FROM JobCategories ORDER BY CategoryName")
        categories_for_dropdown = cursor.fetchall()
    except Exception as e:
        current_app.logger.error(f"Error fetching submissions for review: {e}", exc_info=True)
        flash("Could not load submissions for review.", "danger")
    finally:
        if conn and conn.is_connected(): 
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()
    return render_template('agency_staff_portal/job_offers/review_client_submissions.html', title='Review Client Submissions', submissions=submissions, categories_for_dropdown=categories_for_dropdown)

@job_offer_mgmt_bp.route('/review-client-submission/<int:submission_id>/action', methods=['POST'])
@login_required_with_role(JOB_OFFER_REVIEW_ROLES)
def review_client_submission_action(submission_id):
    action = request.form.get('action')
    reviewer_comments = request.form.get('reviewer_comments', '').strip()
    category_id_on_approval = request.form.get('category_id_on_approval')

    # Basic validation
    if action not in ['approve', 'reject', 'needs_clarification']:
        flash('Invalid action.', 'danger')
        return redirect(url_for('.list_review_client_submissions'))
    if action in ['reject', 'needs_clarification'] and not reviewer_comments:
        flash('Comments are required when rejecting or requesting clarification.', 'warning')
        return redirect(url_for('.list_review_client_submissions'))
    if action == 'approve' and not category_id_on_approval:
        flash('A job category must be selected to approve an offer.', 'warning')
        return redirect(url_for('.list_review_client_submissions'))

    conn = None
    try:
        conn = get_db_connection()
        conn.start_transaction()
        cursor = conn.cursor(dictionary=True)
        
        # Lock the submission row to prevent race conditions
        cursor.execute("SELECT csjo.*, c.CompanyName FROM ClientSubmittedJobOffers csjo JOIN Companies c ON csjo.CompanyID = c.CompanyID WHERE csjo.SubmissionID = %s FOR UPDATE", (submission_id,))
        sub = cursor.fetchone()
        
        if not sub or sub['ReviewStatus'] != 'Pending':
            flash('Submission not found or has already been processed.', 'danger')
            conn.rollback()
            return redirect(url_for('.list_review_client_submissions'))

        if action == 'approve':
            # This is the critical mapping from the submission table to the live offers table.
            live_offer_params = {
                "CompanyID": sub.get('CompanyID'),
                "PostedByStaffID": current_user.specific_role_id,
                "CategoryID": category_id_on_approval,
                "Title": sub.get('Title'),
                "Location": sub.get('Location'),
                "NetSalary": sub.get('NetSalary'),
                "PaymentTerm": sub.get('PaymentTerm'),
                "HiringPlan": sub.get('HiringPlan'),
                "MaxAge": sub.get('MaxAge'),
                "HasContract": sub.get('HasContract'),
                "LanguagesType": sub.get('LanguagesType'),
                "RequiredLanguages": sub.get('RequiredLanguages'),
                "RequiredLevel": sub.get('RequiredLevel'),
                "GraduationStatusRequirement": sub.get('GraduationStatusRequirement'),
                "CandidatesNeeded": sub.get('CandidatesNeeded'),
                "HiringCadence": sub.get('HiringCadence'),
                "WorkLocationType": sub.get('WorkLocationType'),
                "ShiftType": sub.get('ShiftType'),
                "AvailableShifts": sub.get('AvailableShifts'),
                "BenefitsIncluded": sub.get('BenefitsIncluded'),
                "InterviewType": sub.get('InterviewType'),
                "Nationality": sub.get('Nationality'),
                "Gender": sub.get('Gender'),
                "MilitaryStatus": sub.get('MilitaryStatus'),
                "ClosingDate": sub.get('ClosingDate')
                # Note: WorkingDays, WorkingHours, ExperienceRequirement are not on the client form
            }

            sql_live = """INSERT INTO JobOffers (
                            CompanyID, PostedByStaffID, CategoryID, Title, Location, NetSalary, PaymentTerm, HiringPlan,
                            MaxAge, HasContract, LanguagesType, RequiredLanguages, RequiredLevel, GraduationStatusRequirement,
                            CandidatesNeeded, HiringCadence, WorkLocationType, ShiftType, AvailableShifts,
                            BenefitsIncluded, InterviewType, Nationality, Gender, MilitaryStatus, ClosingDate, 
                            Status, DatePosted
                          ) VALUES (
                            %(CompanyID)s, %(PostedByStaffID)s, %(CategoryID)s, %(Title)s, %(Location)s, %(NetSalary)s, %(PaymentTerm)s, %(HiringPlan)s,
                            %(MaxAge)s, %(HasContract)s, %(LanguagesType)s, %(RequiredLanguages)s, %(RequiredLevel)s, %(GraduationStatusRequirement)s,
                            %(CandidatesNeeded)s, %(HiringCadence)s, %(WorkLocationType)s, %(ShiftType)s, %(AvailableShifts)s,
                            %(BenefitsIncluded)s, %(InterviewType)s, %(Nationality)s, %(Gender)s, %(MilitaryStatus)s, %(ClosingDate)s,
                            'Open', NOW()
                          )"""
            cursor.execute(sql_live, live_offer_params)
            live_offer_id = cursor.lastrowid
            
            # Update the submission record
            cursor.execute("""
                UPDATE ClientSubmittedJobOffers 
                SET ReviewStatus='Approved', ReviewerUserID=%s, ReviewDate=NOW(), ReviewerComments=%s, CorrespondingOfferID=%s 
                WHERE SubmissionID=%s
            """, (current_user.id, reviewer_comments, live_offer_id, submission_id))
            
            # Create an announcement
            _create_automated_announcement(
                cursor, 'Automated_Offer', f"New Job: {sub['Title']} with {sub['CompanyName']}", 
                "A client-submitted job offer has been approved and posted.", posted_by_user_id=current_user.id
            )
            flash('Submission approved and job offer has been posted live.', 'success')
        
        else: # Handle reject and needs_clarification
            new_status = 'Rejected' if action == 'reject' else 'NeedsClarification'
            cursor.execute("""
                UPDATE ClientSubmittedJobOffers 
                SET ReviewStatus=%s, ReviewerUserID=%s, ReviewDate=NOW(), ReviewerComments=%s 
                WHERE SubmissionID=%s
            """, (new_status, current_user.id, reviewer_comments, submission_id))
            flash(f'Submission has been marked as "{new_status}".', 'info')
        
        conn.commit()
    except Exception as e:
        if conn: conn.rollback()
        current_app.logger.error(f"Error processing submission {submission_id}, action {action}: {e}", exc_info=True)
        flash(f'An error occurred: {e}', 'danger')
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()
    return redirect(url_for('.list_review_client_submissions'))

@job_offer_mgmt_bp.route('/view-live/<int:offer_id>')
@login_required_with_role(EXECUTIVE_ROLES)
def view_live_job_offer_detail(offer_id):
    offer, conn = None, None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT jo.*, c.CompanyName, jc.CategoryName, 
                   poster_user.FirstName as PosterFirstName, poster_user.LastName as PosterLastName
            FROM JobOffers jo
            JOIN Companies c ON jo.CompanyID = c.CompanyID
            LEFT JOIN JobCategories jc ON jo.CategoryID = jc.CategoryID
            LEFT JOIN Staff poster_staff ON jo.PostedByStaffID = poster_staff.StaffID
            LEFT JOIN Users poster_user ON poster_staff.UserID = poster_user.UserID
            WHERE jo.OfferID = %s
        """, (offer_id,))
        offer = cursor.fetchone()
        
        if offer: 
            def parse_set_column(value):
                if isinstance(value, str): return [v.strip() for v in value.split(',') if v.strip()]
                return []
                
            offer['Benefits_list'] = parse_set_column(offer.get('BenefitsIncluded'))
            offer['AvailableShifts_list'] = parse_set_column(offer.get('AvailableShifts'))
            offer['RequiredLanguages_list'] = parse_set_column(offer.get('RequiredLanguages'))
            offer['GraduationStatus_list'] = parse_set_column(offer.get('GraduationStatusRequirement'))

    except Exception as e:
        current_app.logger.error(f"Error fetching live offer detail {offer_id}: {e}", exc_info=True)
        flash("Could not load offer details.", "danger")
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals(): cursor.close()
            conn.close()
            
    if not offer: 
        flash('Job offer not found.', 'danger')
        return redirect(url_for('.list_all_job_offers'))
        
    return render_template('agency_staff_portal/job_offers/view_live_job_offer_detail.html', title=f"Job Offer Details", offer=offer)


@job_offer_mgmt_bp.route('/applications')
@login_required_with_role(EXECUTIVE_ROLES)
def list_all_applications():
    """
    Unified page to list all job applications with filtering capabilities
    by company, offer, and status.
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    selected_company_id = request.args.get('company_id', type=int)
    selected_offer_id = request.args.get('offer_id', type=int)
    filter_status = request.args.get('filter_status')
    export_format = request.args.get('format')

    cursor.execute("SELECT CompanyID, CompanyName FROM Companies ORDER BY CompanyName")
    companies = cursor.fetchall()
    cursor.execute("SELECT OfferID, Title FROM JobOffers ORDER BY Title")
    job_offers = cursor.fetchall()
    
    sql = """
        SELECT
            u.FirstName, u.LastName, ja.ApplicationDate, jo.Title AS JobTitle,
            c.CompanyName, ja.Status, cand.CandidateID, ja.ApplicationID
        FROM JobApplications ja
        JOIN Candidates cand ON ja.CandidateID = cand.CandidateID
        JOIN Users u ON cand.UserID = u.UserID
        JOIN JobOffers jo ON ja.OfferID = jo.OfferID
        JOIN Companies c ON jo.CompanyID = c.CompanyID
    """
    
    conditions = []
    params = []
    if selected_company_id:
        conditions.append("c.CompanyID = %s")
        params.append(selected_company_id)
    if selected_offer_id:
        conditions.append("jo.OfferID = %s")
        params.append(selected_offer_id)
    if filter_status:
        conditions.append("ja.Status = %s")
        params.append(filter_status)

    if conditions:
        sql += " WHERE " + " AND ".join(conditions)
    
    sql += " ORDER BY ja.ApplicationDate DESC"
    
    cursor.execute(sql, tuple(params))
    applications = cursor.fetchall()
    cursor.close()
    conn.close()

    # --- FIX: FULL EXPORT LOGIC IS NOW IMPLEMENTED ---
    if export_format:
        if not applications:
            flash("No data to export for the selected filters.", "warning")
            return redirect(url_for('.list_all_applications', 
                                    company_id=selected_company_id, 
                                    offer_id=selected_offer_id, 
                                    filter_status=filter_status))
        
        # --- XLSX Export Logic ---
        if export_format == 'xlsx':
            header_map = {
                'Candidate Name': 'FullName',
                'Application Date': 'ApplicationDate',
                'Job Title': 'JobTitle',
                'Company': 'CompanyName',
                'Status': 'Status'
            }
            excel_data = []
            for app in applications:
                new_app = app.copy()
                new_app['FullName'] = f"{app['FirstName']} {app['LastName']}"
                excel_data.append(new_app)
            
            excel_file = _create_styled_excel(excel_data, "All Job Applications", header_map)
            response = make_response(excel_file.read())
            filename = f"all_applications_{datetime.date.today()}.xlsx"
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return response

        # --- CSV Export Logic ---
        if export_format == 'csv':
            output = io.StringIO()
            # Prepare data with a 'Candidate Name' field
            csv_data = []
            if not applications: # Should be caught above, but safe check
                writer = csv.writer(output)
                writer.writerow(['No Data Found'])
            else:
                fieldnames = ['Candidate Name', 'Application Date', 'Job Title', 'Company', 'Status']
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for app in applications:
                    writer.writerow({
                        'Candidate Name': f"{app['FirstName']} {app['LastName']}",
                        'Application Date': app['ApplicationDate'].strftime('%Y-%m-%d') if app['ApplicationDate'] else '',
                        'Job Title': app['JobTitle'],
                        'Company': app['CompanyName'],
                        'Status': app['Status']
                    })
            
            output.seek(0)
            response = make_response(output.read())
            response.headers["Content-Disposition"] = f"attachment; filename=all_applications_{datetime.date.today()}.csv"
            response.headers["Content-type"] = "text/csv"
            return response

    # This only runs if 'export_format' is not provided
    return render_template('agency_staff_portal/job_offers/list_all_applications.html',
                           title="Job Applications",
                           applications=applications,
                           companies=companies,
                           job_offers=job_offers,
                           selected_company_id=selected_company_id,
                           selected_offer_id=selected_offer_id,
                           selected_filter_status=filter_status)

@job_offer_mgmt_bp.route('/company/<int:company_id>/schedules', methods=['GET'])
@login_required_with_role(EXECUTIVE_ROLES)
def view_company_schedules(company_id):
    """
    Displays the form for managing a company's interview schedule. (GET)
    """
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT CompanyName FROM Companies WHERE CompanyID = %s", (company_id,))
        company = cursor.fetchone()
        if not company:
            flash("Company not found.", "danger")
            return redirect(url_for('.list_all_job_offers'))

        cursor.execute("""
            SELECT ScheduleID, DayOfWeek, StartTime, EndTime 
            FROM CompanyInterviewSchedules 
            WHERE CompanyID = %s
            ORDER BY FIELD(DayOfWeek, 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'), StartTime
        """, (company_id,))
        schedules = cursor.fetchall()

        day_options = get_column_options(conn, cursor, 'CompanyInterviewSchedules', 'DayOfWeek')

    except Exception as e:
        current_app.logger.error(f"Error viewing schedules for company {company_id}: {e}", exc_info=True)
        flash("An error occurred while loading the schedule page.", "danger")
        return redirect(url_for('.list_all_job_offers'))
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()

    return render_template('agency_staff_portal/job_offers/manage_company_schedules.html',
                           title=f"Interview Schedule for {company['CompanyName']}",
                           company=company,
                           schedules=schedules,
                           day_options=day_options,
                           company_id=company_id)


@job_offer_mgmt_bp.route('/company/<int:company_id>/schedules/add', methods=['POST'])
@login_required_with_role(EXECUTIVE_ROLES)
def add_schedule_slot(company_id):
    """
    Adds a single new interview schedule slot for a company. (POST)
    """
    day = request.form.get('day')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')

    if not all([day, start_time, end_time]):
        flash("All fields (Day, Start Time, End Time) are required to add a new slot.", "warning")
        return redirect(url_for('.view_company_schedules', company_id=company_id))

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        insert_sql = """
            INSERT INTO CompanyInterviewSchedules (CompanyID, DayOfWeek, StartTime, EndTime)
            VALUES (%s, %s, %s, %s)
        """
        cursor.execute(insert_sql, (company_id, day, start_time, end_time))
        conn.commit()
        flash("New schedule slot added successfully!", "success")
    except Exception as e:
        if conn and conn.is_connected(): conn.rollback()
        current_app.logger.error(f"Error adding schedule for company {company_id}: {e}", exc_info=True)
        flash("An error occurred while adding the schedule slot.", "danger")
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()
            
    return redirect(url_for('.view_company_schedules', company_id=company_id))


@job_offer_mgmt_bp.route('/schedules/delete/<int:schedule_id>', methods=['POST'])
@login_required_with_role(EXECUTIVE_ROLES)
def delete_schedule_slot(schedule_id):
    """
    Deletes a single interview schedule slot. (POST)
    """
    conn = get_db_connection()
    company_id = None
    try:
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT CompanyID FROM CompanyInterviewSchedules WHERE ScheduleID = %s", (schedule_id,))
        result = cursor.fetchone()
        if result:
            company_id = result['CompanyID']
        
        cursor.execute("DELETE FROM CompanyInterviewSchedules WHERE ScheduleID = %s", (schedule_id,))
        conn.commit()
        
        if cursor.rowcount > 0:
            flash("Schedule slot deleted successfully.", "success")
        else:
            flash("Schedule slot not found.", "warning")

    except Exception as e:
        if conn and conn.is_connected(): conn.rollback()
        current_app.logger.error(f"Error deleting schedule slot {schedule_id}: {e}", exc_info=True)
        flash("An error occurred while deleting the schedule slot.", "danger")
        if not company_id:
            return redirect(url_for('.list_all_job_offers'))
    finally:
        if conn and conn.is_connected():
            if 'cursor' in locals() and cursor: cursor.close()
            conn.close()

    return redirect(url_for('.view_company_schedules', company_id=company_id))

@job_offer_mgmt_bp.route('/applications/view/<int:application_id>')
@login_required_with_role(EXECUTIVE_ROLES)
def view_application_details(application_id):
    """
    Displays a comprehensive view of a single job application, including
    candidate profile, interview details, and notes.
    """
    conn = get_db_connection()
    details = {}
    try:
        cursor = conn.cursor(dictionary=True)
        sql = """
            SELECT
                ja.ApplicationID, ja.ApplicationDate, ja.Status AS ApplicationStatus,
                ja.NotesByCandidate, ja.NotesByStaff,
                u.FirstName, u.LastName, u.Email, u.PhoneNumber, u.ProfilePictureURL,
                cand.CandidateID, cand.DateOfBirth, cand.Nationality, cand.EducationalStatus, cand.Languages, cand.LanguageLevel,
                jo.OfferID, jo.Title AS JobTitle,
                c.CompanyID, c.CompanyName,
                i.InterviewID, i.ScheduledDateTime, i.Status AS InterviewStatus, i.Notes AS InterviewNotes
            FROM JobApplications ja
            JOIN Candidates cand ON ja.CandidateID = cand.CandidateID
            JOIN Users u ON cand.UserID = u.UserID
            JOIN JobOffers jo ON ja.OfferID = jo.OfferID
            JOIN Companies c ON jo.CompanyID = c.CompanyID
            LEFT JOIN Interviews i ON ja.ApplicationID = i.ApplicationID
            WHERE ja.ApplicationID = %s
        """
        cursor.execute(sql, (application_id,))
        details = cursor.fetchone()

        if not details:
            flash("Application not found.", "danger")
            return redirect(url_for('.list_all_applications'))

    except Exception as e:
        current_app.logger.error(f"Error fetching application details for ID {application_id}: {e}", exc_info=True)
        flash("Could not load application details.", "danger")
        return redirect(url_for('.list_all_applications'))
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

    return render_template('agency_staff_portal/job_offers/view_application_detail.html',
                           title=f"Reviewing {details['FirstName']} for {details['JobTitle']}",
                           details=details)