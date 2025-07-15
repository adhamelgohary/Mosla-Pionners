# routes/Agency_staff_portal/reporting_routes.py

from flask import Blueprint, render_template, request, flash, url_for, make_response
from utils.decorators import login_required_with_role, MANAGERIAL_PORTAL_ROLES
from db import get_db_connection
import datetime
import csv
import io

# --- New Imports for Excel Generation ---
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

reporting_bp = Blueprint('reporting_bp', __name__,
                         template_folder='../../../templates',
                         url_prefix='/managerial/reports')

# --- Local Helper Function for Excel Styling (No separate file needed) ---
def _create_styled_excel(report_data, title, header_mapping):
    """
    Generates a styled Excel report using only OpenPyXL.

    Args:
        report_data (list): A list of dictionaries with the data.
        title (str): The main title for the report.
        header_mapping (dict): A dictionary mapping { 'Excel Header Name': 'data_key' }.

    Returns:
        io.BytesIO: An in-memory binary stream of the Excel file.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    # --- Define Styles ---
    title_font = Font(name='Calibri', size=18, bold=True, color='1F2937')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

    # --- Add and Style Title & Subtitle ---
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header_mapping))
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header_mapping))
    subtitle_cell = worksheet.cell(row=2, column=1, value=f"Report generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    subtitle_cell.font = Font(italic=True, color='6B7280')
    subtitle_cell.alignment = center_align
    worksheet.row_dimensions[2].height = 20

    # --- Write and Style Headers ---
    headers = list(header_mapping.keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    worksheet.row_dimensions[4].height = 20

    # --- Write Data Cell by Cell ---
    data_keys = list(header_mapping.values())
    for row_num, row_data in enumerate(report_data, 5):
        for col_num, key in enumerate(data_keys, 1):
            cell_value = row_data.get(key, 'N/A')
            # Format dates nicely if they are datetime objects
            if isinstance(cell_value, (datetime.datetime, datetime.date)):
                cell_value = cell_value.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=col_num, value=cell_value).alignment = left_align

    # --- Adjust Column Widths ---
    column_widths = {}
    for col_num, header_title in enumerate(headers, 1):
        column_widths[col_num] = len(header_title)

    for row_data in report_data:
        for col_num, key in enumerate(data_keys, 1):
            cell_len = len(str(row_data.get(key, '')))
            if cell_len > column_widths[col_num]:
                column_widths[col_num] = cell_len
    
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(col_num)].width = width + 4


    # --- Save to an in-memory stream ---
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

@reporting_bp.route('/')
@login_required_with_role(MANAGERIAL_PORTAL_ROLES)
def reporting_hub():
    """Main dashboard for all available reports."""
    return render_template('agency_staff_portal/reports/reporting_hub.html', title="Reporting Hub")


@reporting_bp.route('/hiring-performance', methods=['GET'])
@login_required_with_role(MANAGERIAL_PORTAL_ROLES)
def hiring_performance_report():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    params = []
    sql = """
        SELECT 
            jo.Title,
            c.CompanyName,
            jo.Status,
            jo.DatePosted,
            (SELECT COUNT(*) FROM JobApplications ja WHERE ja.OfferID = jo.OfferID) as TotalApplicants,
            DATEDIFF(jo.FilledDate, jo.DatePosted) as TimeToFill
        FROM JobOffers jo
        JOIN Companies c ON jo.CompanyID = c.CompanyID
    """
    
    conditions = []
    if start_date_str:
        conditions.append("jo.DatePosted >= %s")
        params.append(start_date_str)
    if end_date_str:
        conditions.append("jo.DatePosted <= %s")
        params.append(end_date_str)

    if conditions:
        sql += " WHERE " + " AND ".join(conditions)
    
    sql += " ORDER BY jo.DatePosted DESC"
    
    cursor.execute(sql, tuple(params))
    report_data = cursor.fetchall()
    cursor.close()
    conn.close()

    export_format = request.args.get('format')

    if export_format and not report_data:
        flash("No data to export for the selected filters.", "warning")
        return render_template('agency_staff_portal/reports/hiring_performance.html',
                           title="Hiring Performance Report",
                           report_data=[], start_date=start_date_str, end_date=end_date_str)

    if export_format == 'xlsx':
        header_map = {
            'Job Title': 'Title',
            'Company': 'CompanyName',
            'Status': 'Status',
            'Date Posted': 'DatePosted',
            'Applicants': 'TotalApplicants',
            'Time to Fill (Days)': 'TimeToFill'
        }
        excel_file = _create_styled_excel(report_data, "Hiring Performance Report", header_map)
        
        response = make_response(excel_file.read())
        filename = f"hiring_report_{datetime.date.today()}.xlsx"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    if export_format == 'csv':
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=report_data[0].keys())
        writer.writeheader()
        writer.writerows(report_data)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers["Content-Disposition"] = f"attachment; filename=hiring_report_{datetime.date.today()}.csv"
        response.headers["Content-type"] = "text/csv"
        return response

    return render_template('agency_staff_portal/reports/hiring_performance.html',
                           title="Hiring Performance Report",
                           report_data=report_data,
                           start_date=start_date_str,
                           end_date=end_date_str)


@reporting_bp.route('/staff-performance', methods=['GET'])
@login_required_with_role(MANAGERIAL_PORTAL_ROLES)
def staff_performance_report():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    selected_role = request.args.get('role')

    cursor.execute("SELECT DISTINCT Role FROM Staff ORDER BY Role")
    roles = [row['Role'] for row in cursor.fetchall()]

    params = []
    date_conditions = []
    if start_date_str: date_conditions.append("ja.ApplicationDate >= %s"); params.append(start_date_str)
    if end_date_str: date_conditions.append("ja.ApplicationDate <= %s"); params.append(end_date_str)
    app_date_filter = "WHERE " + " AND ".join(date_conditions) if date_conditions else ""

    offer_params = []
    offer_date_conditions = []
    if start_date_str: offer_date_conditions.append("jo.DatePosted >= %s"); offer_params.append(start_date_str)
    if end_date_str: offer_date_conditions.append("jo.DatePosted <= %s"); offer_params.append(end_date_str)
    offer_date_filter = "WHERE " + " AND ".join(offer_date_conditions) if offer_date_conditions else ""
    
    points_params = []
    points_date_conditions = []
    if start_date_str: points_date_conditions.append("spl.AwardDate >= %s"); points_params.append(start_date_str)
    if end_date_str: points_date_conditions.append("spl.AwardDate <= %s"); points_params.append(end_date_str)
    points_date_filter = "WHERE " + " AND ".join(points_date_conditions) if points_date_conditions else ""
    
    sql_params = params + params + offer_params + points_params

    sql = f"""
        SELECT
            s.StaffID,
            CONCAT(u.FirstName, ' ', u.LastName) as StaffName,
            s.Role,
            COALESCE(ja_referred.TotalReferred, 0) as TotalApplicationsReferred,
            COALESCE(ja_hired.TotalHired, 0) as TotalHires,
            COALESCE(jo_posted.TotalOffersPosted, 0) as TotalOffersPosted,
            COALESCE(spl_points.TotalPoints, 0) as TotalPoints
        FROM Staff s
        JOIN Users u ON s.UserID = u.UserID
        LEFT JOIN (
            SELECT ReferringStaffID, COUNT(ApplicationID) as TotalReferred FROM JobApplications ja {app_date_filter} GROUP BY ReferringStaffID
        ) as ja_referred ON s.StaffID = ja_referred.ReferringStaffID
        LEFT JOIN (
            SELECT ReferringStaffID, COUNT(ApplicationID) as TotalHired FROM JobApplications ja WHERE ja.Status = 'Hired' {'AND ' + ' AND '.join(date_conditions) if date_conditions else ''} GROUP BY ReferringStaffID
        ) as ja_hired ON s.StaffID = ja_hired.ReferringStaffID
        LEFT JOIN (
            SELECT PostedByStaffID, COUNT(OfferID) as TotalOffersPosted FROM JobOffers jo {offer_date_filter} GROUP BY PostedByStaffID
        ) as jo_posted ON s.StaffID = jo_posted.PostedByStaffID
        LEFT JOIN (
            SELECT AwardedToStaffID, SUM(PointsAmount) as TotalPoints FROM StaffPointsLog spl {points_date_filter} GROUP BY AwardedToStaffID
        ) as spl_points ON s.StaffID = spl_points.AwardedToStaffID
    """
    
    role_condition = ""
    if selected_role:
        role_condition = "WHERE s.Role = %s"
        sql_params.append(selected_role)
    
    sql += f" {role_condition} HAVING TotalApplicationsReferred > 0 OR TotalOffersPosted > 0 OR TotalPoints > 0 ORDER BY StaffName;"

    cursor.execute(sql, tuple(sql_params))
    report_data = cursor.fetchall()
    cursor.close()
    conn.close()
    
    export_format = request.args.get('format')
    
    if export_format and not report_data:
        flash("No data to export for the selected filters.", "warning")
        return render_template('agency_staff_portal/reports/staff_performance.html',
                           title="Staff Performance Report",
                           report_data=[], roles=roles, start_date=start_date_str,
                           end_date=end_date_str, selected_role=selected_role)

    if export_format == 'xlsx':
        excel_data = []
        for row in report_data:
            new_row = row.copy()
            if new_row.get('TotalApplicationsReferred', 0) > 0:
                hire_rate = (new_row.get('TotalHires', 0) / new_row['TotalApplicationsReferred']) * 100
                new_row['HireRate'] = f"{hire_rate:.2f}%"
            else:
                new_row['HireRate'] = "0.00%"
            excel_data.append(new_row)
        
        header_map = {
            'Staff Name': 'StaffName',
            'Role': 'Role',
            'Apps Referred': 'TotalApplicationsReferred',
            'Hires': 'TotalHires',
            'Hire Rate': 'HireRate',
            'Offers Posted': 'TotalOffersPosted',
            'Points Earned': 'TotalPoints'
        }
        excel_file = _create_styled_excel(excel_data, "Staff Performance Report", header_map)
        
        response = make_response(excel_file.read())
        filename = f"staff_performance_report_{datetime.date.today()}.xlsx"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    if export_format == 'csv':
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=report_data[0].keys())
        writer.writeheader()
        writer.writerows(report_data)
        output.seek(0)
        response = make_response(output.read())
        response.headers["Content-Disposition"] = f"attachment; filename=staff_performance_report_{datetime.date.today()}.csv"
        response.headers["Content-type"] = "text/csv"
        return response

    return render_template('agency_staff_portal/reports/staff_performance.html',
                           title="Staff Performance Report",
                           report_data=report_data,
                           roles=roles,
                           start_date=start_date_str,
                           end_date=end_date_str,
                           selected_role=selected_role)