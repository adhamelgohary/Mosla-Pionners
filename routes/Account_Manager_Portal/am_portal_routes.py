# routes/Account_Manager_Portal/am_portal_routes.py
import datetime
import io
import csv

from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, abort, make_response
from flask_login import current_user
from utils.decorators import login_required_with_role
from db import get_db_connection

# --- Excel Styling Imports ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- Roles are defined once for clarity, matching the new schema ---
AM_PORTAL_ACCESS_ROLES = ['AccountManager', 'SeniorAccountManager', 'HeadAccountManager', 'CEO', 'Founder']
STAFF_MANAGEMENT_ROLES = ['HeadAccountManager', 'CEO', 'Founder']
MANAGEABLE_STAFF_ROLES = ['AccountManager', 'SeniorAccountManager']

account_manager_bp = Blueprint('account_manager_bp', __name__,
                               template_folder='../../../templates',
                               url_prefix='/am-portal')

# --- NEW: Reusable Styled Excel Helper Function ---
def _create_styled_excel(report_data, title, header_mapping):
    """
    Generates a styled Excel report using only OpenPyXL, matching the managerial portal style.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    # --- Define Styles (Consistent with other reports) ---
    title_font = Font(name='Calibri', size=18, bold=True, color='1F2937')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo

    # --- Add and Style Title & Subtitle ---
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header_mapping))
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header_mapping))
    subtitle_cell = worksheet.cell(row=2, column=1, value=f"Report generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    subtitle_cell.font = Font(italic=True, color='6B7280')
    subtitle_cell.alignment = center_align
    worksheet.row_dimensions[2].height = 20

    # --- Write and Style Headers ---
    headers = list(header_mapping.keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    worksheet.row_dimensions[4].height = 20

    # --- Write Data Cell by Cell ---
    data_keys = list(header_mapping.values())
    for row_num, row_data in enumerate(report_data, 5):
        for col_num, key in enumerate(data_keys, 1):
            cell_value = row_data.get(key, 'N/A')
            if isinstance(cell_value, (datetime.datetime, datetime.date)):
                cell_value = cell_value.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=col_num, value=cell_value).alignment = left_align

    # --- Adjust Column Widths ---
    column_widths = {}
    for col_num, header_title in enumerate(headers, 1):
        column_widths[col_num] = len(str(header_title))
    for row_data in report_data:
        for col_num, key in enumerate(data_keys, 1):
            cell_len = len(str(row_data.get(key, '')))
            if cell_len > column_widths.get(col_num, 0):
                column_widths[col_num] = cell_len
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(col_num)].width = width + 4

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def _is_user_authorized_for_application(staff_id, application_id):
    """Checks if a staff member can action a specific application, including hierarchy."""
    if not staff_id or not application_id: return False
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT c.ManagedByStaffID FROM JobApplications ja
            JOIN JobOffers jo ON ja.OfferID = jo.OfferID
            JOIN Companies c ON jo.CompanyID = c.CompanyID
            WHERE ja.ApplicationID = %s
        """, (application_id,))
        result = cursor.fetchone()
        if not result or not result['ManagedByStaffID']: return False
        
        company_manager_staff_id = result['ManagedByStaffID']
        if company_manager_staff_id == staff_id: return True

        current_id_in_chain = company_manager_staff_id
        for _ in range(5): # Limit recursion depth for safety
            cursor.execute("SELECT ReportsToStaffID FROM Staff WHERE StaffID = %s", (current_id_in_chain,))
            supervisor = cursor.fetchone()
            if not supervisor or not supervisor['ReportsToStaffID']: return False
            current_id_in_chain = supervisor['ReportsToStaffID']
            if current_id_in_chain == staff_id: return True
        return False
    finally:
        if conn and conn.is_connected():
             cursor.close()
             conn.close()

@account_manager_bp.route('/')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def portal_home():
    return redirect(url_for('.dashboard'))

@account_manager_bp.route('/dashboard')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def dashboard():
    staff_id = getattr(current_user, 'specific_role_id', None)
    if not staff_id:
        flash("Your staff profile ID could not be found.", "danger")
        return redirect(url_for('staff_dashboard_bp.main_dashboard'))

    conn = get_db_connection()
    dashboard_data = {}
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as count FROM Companies WHERE ManagedByStaffID = %s", (staff_id,))
        dashboard_data['managed_companies_count'] = cursor.fetchone()['count']
        cursor.execute("SELECT COUNT(*) as count FROM JobOffers jo JOIN Companies c ON jo.CompanyID = c.CompanyID WHERE c.ManagedByStaffID = %s AND jo.Status = 'Open'", (staff_id,))
        dashboard_data['open_offers_count'] = cursor.fetchone()['count']
        cursor.execute("SELECT COUNT(ja.ApplicationID) as count FROM JobApplications ja JOIN JobOffers jo ON ja.OfferID = jo.OfferID JOIN Companies c ON jo.CompanyID = c.CompanyID WHERE c.ManagedByStaffID = %s AND ja.Status = 'Shortlisted'", (staff_id,))
        dashboard_data['pending_interview_scheduling_count'] = cursor.fetchone()['count']
        cursor.execute("SELECT c.CompanyID, c.CompanyName, c.CompanyLogoURL, (SELECT COUNT(*) FROM JobOffers WHERE CompanyID = c.CompanyID AND Status = 'Open') as OpenJobs FROM Companies c WHERE ManagedByStaffID = %s ORDER BY CompanyName", (staff_id,))
        dashboard_data['managed_companies_list'] = cursor.fetchall()
        cursor.execute("SELECT u.FirstName, u.LastName, jo.Title as JobTitle, ja.ApplicationDate, ja.ApplicationID, comp.CompanyName FROM JobApplications ja JOIN Candidates c ON ja.CandidateID = c.CandidateID JOIN Users u ON c.UserID = u.UserID JOIN JobOffers jo ON ja.OfferID = jo.OfferID JOIN Companies comp ON jo.CompanyID = comp.CompanyID WHERE comp.ManagedByStaffID = %s AND ja.Status IN ('Applied', 'Submitted') ORDER BY ja.ApplicationDate DESC LIMIT 5", (staff_id,))
        dashboard_data['recent_applicants'] = cursor.fetchall()
    finally:
        if conn and conn.is_connected(): conn.close()
            
    return render_template('account_manager_portal/dashboard.html', title="Account Manager Dashboard", dashboard_data=dashboard_data)


@account_manager_bp.route('/interview-pipeline')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def interview_pipeline():
    staff_id = getattr(current_user, 'specific_role_id', None)
    if not staff_id:
        flash("Your staff profile ID could not be found.", "danger")
        return redirect(url_for('.dashboard'))
    conn = get_db_connection()
    pipeline_apps = []
    companies_with_scheduled_interviews = []
    try:
        cursor = conn.cursor(dictionary=True)
        # Main query to get all pipeline applications
        sql = """
            SELECT ja.ApplicationID, ja.Status, u.FirstName, u.LastName, u.Email, 
                   jo.Title as OfferTitle, comp.CompanyName, comp.CompanyID,
                   i.InterviewID, i.ScheduledDateTime, i.Status as InterviewStatus
            FROM JobApplications ja
            JOIN Candidates c ON ja.CandidateID = c.CandidateID
            JOIN Users u ON c.UserID = u.UserID
            JOIN JobOffers jo ON ja.OfferID = jo.OfferID
            JOIN Companies comp ON jo.CompanyID = comp.CompanyID
            LEFT JOIN Interviews i ON ja.ApplicationID = i.ApplicationID
            WHERE ja.Status IN ('Shortlisted', 'Interview Scheduled') AND comp.ManagedByStaffID = %s
            ORDER BY CASE WHEN ja.Status = 'Shortlisted' THEN 0 ELSE 1 END, i.ScheduledDateTime, ja.ApplicationDate DESC;
        """
        cursor.execute(sql, (staff_id,))
        pipeline_apps = cursor.fetchall()
        
        # New query to get a unique list of companies for the filter dropdown
        if pipeline_apps:
            company_filter_sql = """
                SELECT DISTINCT comp.CompanyID, comp.CompanyName
                FROM JobApplications ja
                JOIN JobOffers jo ON ja.OfferID = jo.OfferID
                JOIN Companies comp ON jo.CompanyID = comp.CompanyID
                WHERE ja.Status = 'Interview Scheduled' AND comp.ManagedByStaffID = %s
                ORDER BY comp.CompanyName;
            """
            cursor.execute(company_filter_sql, (staff_id,))
            companies_with_scheduled_interviews = cursor.fetchall()

    finally:
        if conn and conn.is_connected(): conn.close()
            
    return render_template('account_manager_portal/interview_pipeline.html', 
                           title="Interview Pipeline", 
                           pipeline_apps=pipeline_apps,
                           companies_with_scheduled_interviews=companies_with_scheduled_interviews)
    
    
@account_manager_bp.route('/my-staff')
@login_required_with_role(STAFF_MANAGEMENT_ROLES)
def my_staff():
    conn = get_db_connection()
    staff_list = []
    try:
        cursor = conn.cursor(dictionary=True)
        placeholders = ', '.join(['%s'] * len(MANAGEABLE_STAFF_ROLES))
        query = f"SELECT s.StaffID, u.UserID, u.FirstName, u.LastName, s.Role, u.ProfilePictureURL, (SELECT COUNT(*) FROM Companies WHERE ManagedByStaffID = s.StaffID) AS AssignedCompanyCount FROM Staff s JOIN Users u ON s.UserID = u.UserID WHERE s.Role IN ({placeholders}) AND u.IsActive = 1 ORDER BY u.LastName, u.FirstName"
        cursor.execute(query, tuple(MANAGEABLE_STAFF_ROLES))
        staff_list = cursor.fetchall()
    finally:
        if conn and conn.is_connected(): conn.close()
    return render_template('account_manager_portal/my_staff.html', title="My Staff - Account Managers", staff_list=staff_list)

# --- [NEW ROUTE] Master view for Head AMs to see all companies ---
@account_manager_bp.route('/all-companies')
@login_required_with_role(STAFF_MANAGEMENT_ROLES)
def all_companies_overview():
    """
    Provides a master view for senior management to see all companies
    and their assigned managers.
    """
    conn = get_db_connection()
    all_companies_list = []
    try:
        cursor = conn.cursor(dictionary=True)
        query = """
            SELECT 
                c.CompanyID, c.CompanyName, c.Industry,
                s.StaffID AS ManagerStaffID,
                u.FirstName AS ManagerFirstName,
                u.LastName AS ManagerLastName,
                (SELECT COUNT(*) FROM JobOffers jo WHERE jo.CompanyID = c.CompanyID AND jo.Status = 'Open') AS OpenJobsCount
            FROM Companies c
            LEFT JOIN Staff s ON c.ManagedByStaffID = s.StaffID
            LEFT JOIN Users u ON s.UserID = u.UserID
            ORDER BY c.CompanyName;
        """
        cursor.execute(query)
        all_companies_list = cursor.fetchall()
    except Exception as e:
        current_app.logger.error(f"Error fetching all companies overview: {e}", exc_info=True)
        flash("An unexpected error occurred while loading the companies list.", "danger")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

    # NOTE: This route requires a new template: account_manager_portal/all_companies_overview.html
    # This template should display a table of companies with columns for Company Name, Industry,
    # Assigned Manager, Open Jobs, and Action links to view/manage the company.
    return render_template('account_manager_portal/all_companies_overview.html',
                           title="All Companies Overview",
                           companies=all_companies_list)


@account_manager_bp.route('/my-portfolio')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def my_portfolio():
    if not hasattr(current_user, 'specific_role_id'):
        flash("Your staff profile ID could not be found.", "danger")
        return redirect(url_for('.dashboard'))
    return redirect(url_for('.view_manager_portfolio', manager_staff_id=current_user.specific_role_id))

@account_manager_bp.route('/portfolio/<int:manager_staff_id>')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def view_manager_portfolio(manager_staff_id):
    is_own_portfolio = hasattr(current_user, 'specific_role_id') and current_user.specific_role_id == manager_staff_id
    if not is_own_portfolio and current_user.role_type not in ['HeadAccountManager', 'CEO', 'Founder', 'Founder']:
        abort(403, "You are not authorized to view this portfolio.")
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT s.StaffID, u.FirstName, u.LastName, s.Role FROM Staff s JOIN Users u ON s.UserID = u.UserID WHERE s.StaffID = %s", (manager_staff_id,))
        manager_details = cursor.fetchone()
        if not manager_details: abort(404, "Account Manager not found.")
        cursor.execute("SELECT CompanyID, CompanyName, Industry FROM Companies WHERE ManagedByStaffID = %s ORDER BY CompanyName", (manager_staff_id,))
        managed_companies = cursor.fetchall()
        companies_with_data = []
        for company in managed_companies:
            cursor.execute("SELECT OfferID, Title, Status, (SELECT COUNT(*) FROM JobApplications WHERE OfferID = jo.OfferID AND Status IN ('Applied', 'Submitted')) as NewApplicantCount FROM JobOffers jo WHERE CompanyID = %s ORDER BY Status, Title", (company['CompanyID'],))
            company['job_offers'] = cursor.fetchall()
            companies_with_data.append(company)
    finally:
        if conn and conn.is_connected(): conn.close()
    return render_template('account_manager_portal/portfolio_detailed.html', title=f"Portfolio: {manager_details['FirstName']} {manager_details['LastName']}", manager=manager_details, companies_data=companies_with_data)

@account_manager_bp.route('/application/<int:application_id>/update', methods=['POST'])
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def update_application_status(application_id):
    action = request.form.get('action')
    manager_staff_id = request.form.get('manager_staff_id')
    offer_id_for_redirect = request.form.get('offer_id')
    feedback_notes = request.form.get('feedback_notes', '').strip()
    if not all([action, manager_staff_id, offer_id_for_redirect]):
        flash("Invalid request. Missing required data.", "danger")
        return redirect(url_for('.dashboard'))
    if not _is_user_authorized_for_application(current_user.specific_role_id, application_id): abort(403)
    new_status = {'approve': 'Shortlisted', 'reject': 'Rejected'}.get(action)
    if new_status:
        conn = get_db_connection()
        try:
            conn.start_transaction()
            cursor = conn.cursor()
            sql = "UPDATE JobApplications SET Status = %s, NotesByStaff = CONCAT(COALESCE(NotesByStaff, ''), %s) WHERE ApplicationID = %s"
            notes_to_add = f"\n\n--- {new_status} by {current_user.first_name} on {datetime.date.today().strftime('%Y-%m-%d')} ---\n{feedback_notes}"
            cursor.execute(sql, (new_status, notes_to_add, application_id))
            conn.commit()
            flash(f"Application status updated to '{new_status}'.", "success")
        finally:
            if conn and conn.is_connected(): conn.close()
    else: flash("Unknown action specified.", "danger")
    return redirect(url_for('.view_offer_applicants', offer_id=offer_id_for_redirect))

@account_manager_bp.route('/company/<int:company_id>')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def view_single_company(company_id):
    staff_id = getattr(current_user, 'specific_role_id', None)
    if not staff_id: abort(403)
    conn = get_db_connection()
    company_data = {}
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM Companies WHERE CompanyID = %s", (company_id,))
        company_info = cursor.fetchone()
        is_owner = company_info and company_info.get('ManagedByStaffID') == staff_id
        is_senior_manager = current_user.role_type in STAFF_MANAGEMENT_ROLES
        if not (is_owner or is_senior_manager):
            flash("You are not authorized to view this company or it does not exist.", "danger")
            return redirect(url_for('.dashboard'))
        company_data['info'] = company_info
        cursor.execute("SELECT OfferID, Title, Status FROM JobOffers WHERE CompanyID = %s ORDER BY CASE Status WHEN 'Open' THEN 1 ELSE 2 END, Title", (company_id,))
        company_data['job_offers'] = cursor.fetchall()
    finally:
        if conn and conn.is_connected(): conn.close()
    return render_template('account_manager_portal/single_company_view.html', title=f"Manage: {company_data['info']['CompanyName']}", company_data=company_data)

@account_manager_bp.route('/offer/<int:offer_id>/applicants')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def view_offer_applicants(offer_id):
    staff_id = getattr(current_user, 'specific_role_id', None)
    if not staff_id: abort(403)
    conn = get_db_connection()
    offer_data = {}
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT jo.OfferID, jo.Title, jo.Status, c.CompanyID, c.CompanyName, c.ManagedByStaffID FROM JobOffers jo JOIN Companies c ON jo.CompanyID = c.CompanyID WHERE jo.OfferID = %s", (offer_id,))
        offer_info = cursor.fetchone()
        if not offer_info: abort(404, "Job offer not found.")
        is_company_manager = offer_info['ManagedByStaffID'] == staff_id
        is_senior_manager = current_user.role_type in STAFF_MANAGEMENT_ROLES
        if not (is_company_manager or is_senior_manager):
             abort(403, "You are not authorized to view applicants for this offer.")
        offer_data['info'] = offer_info
        cursor.execute("SELECT ja.ApplicationID, ja.Status, ja.ApplicationDate, c.CandidateID, u.UserID, u.FirstName, u.LastName, u.Email, u.ProfilePictureURL FROM JobApplications ja JOIN Candidates c ON ja.CandidateID = c.CandidateID JOIN Users u ON c.UserID = u.UserID WHERE ja.OfferID = %s ORDER BY CASE ja.Status WHEN 'Applied' THEN 1 WHEN 'Submitted' THEN 2 WHEN 'Shortlisted' THEN 3 ELSE 4 END, ja.ApplicationDate DESC", (offer_id,))
        offer_data['applicants'] = cursor.fetchall()
    finally:
        if conn and conn.is_connected(): conn.close()
    return render_template('account_manager_portal/single_offer_view.html', title=f"Applicants for: {offer_data['info']['Title']}", offer_data=offer_data, manager_staff_id=staff_id)
    
@account_manager_bp.route('/application/<int:application_id>/review-details')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def review_application_details(application_id):
    staff_id = getattr(current_user, 'specific_role_id', None)
    if not staff_id or not _is_user_authorized_for_application(staff_id, application_id): abort(403)
    conn = get_db_connection()
    review_data = {}
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT ja.ApplicationID, ja.NotesByCandidate, ja.ApplicationDate, ja.NotesByStaff, c.CandidateID, jo.OfferID, jo.Title as OfferTitle, comp.CompanyName FROM JobApplications ja JOIN Candidates c ON ja.CandidateID = c.CandidateID JOIN JobOffers jo ON ja.OfferID = jo.OfferID JOIN Companies comp ON jo.CompanyID = comp.CompanyID WHERE ja.ApplicationID = %s", (application_id,))
        app_info = cursor.fetchone()
        if not app_info: abort(404, "Application not found.")
        review_data['application'] = app_info
        candidate_id = app_info['CandidateID']
        cursor.execute("SELECT c.*, u.FirstName, u.LastName, u.Email, u.PhoneNumber, u.ProfilePictureURL, u.RegistrationDate FROM Candidates c JOIN Users u ON c.UserID = u.UserID WHERE c.CandidateID = %s", (candidate_id,))
        review_data['candidate_profile'] = cursor.fetchone()
        # Ensure 'Languages' is a list for the template
        if review_data['candidate_profile'] and isinstance(review_data['candidate_profile'].get('Languages'), str):
            review_data['candidate_profile']['Languages'] = review_data['candidate_profile']['Languages'].split(',')

        cursor.execute("SELECT CVID, CVFileUrl, OriginalFileName, CVTitle FROM CandidateCVs WHERE CandidateID = %s AND UploadedAt <= %s ORDER BY IsPrimary DESC, UploadedAt DESC LIMIT 1", (candidate_id, app_info['ApplicationDate']))
        review_data['cv'] = cursor.fetchone()
        cursor.execute("SELECT VoiceNoteID, VoiceNoteURL, Title FROM CandidateVoiceNotes WHERE CandidateID = %s AND UploadedAt <= %s AND Purpose = 'Job Application' ORDER BY UploadedAt DESC LIMIT 1", (candidate_id, app_info['ApplicationDate']))
        review_data['voice_note'] = cursor.fetchone()
    finally:
        if conn and conn.is_connected(): conn.close()
    return render_template('account_manager_portal/application_review_modal.html', review_data=review_data, manager_staff_id=staff_id)


@account_manager_bp.route('/interview-pipeline/export-scheduled')
@login_required_with_role(AM_PORTAL_ACCESS_ROLES)
def export_scheduled_interviews():
    staff_id = getattr(current_user, 'specific_role_id', None)
    if not staff_id:
        abort(403)
        
    company_id_filter = request.args.get('company_id')

    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        sql = """
            SELECT
                u.FirstName, u.LastName, u.Email, u.PhoneNumber, u.RegistrationDate,
                c.LinkedInProfileURL, c.DateOfBirth, c.Nationality AS CandidateNationality, 
                c.Languages, c.LanguageLevel, c.EducationalStatus, c.Gender,
                jo.Title AS OfferTitle,
                comp.CompanyName,
                i.ScheduledDateTime
            FROM JobApplications ja
            JOIN Candidates c ON ja.CandidateID = c.CandidateID
            JOIN Users u ON c.UserID = u.UserID
            JOIN JobOffers jo ON ja.OfferID = jo.OfferID
            JOIN Companies comp ON jo.CompanyID = comp.CompanyID
            JOIN Interviews i ON ja.ApplicationID = i.ApplicationID
            WHERE ja.Status = 'Interview Scheduled' AND comp.ManagedByStaffID = %s
        """
        params = [staff_id]
        if company_id_filter and company_id_filter.isdigit():
            sql += " AND comp.CompanyID = %s"
            params.append(int(company_id_filter))
        sql += " ORDER BY comp.CompanyName, i.ScheduledDateTime;"
        cursor.execute(sql, tuple(params))
        scheduled_interviews = cursor.fetchall()

        if not scheduled_interviews:
            flash("No scheduled interviews found for the selected criteria.", "warning")
            return redirect(url_for('.interview_pipeline'))
        
        # --- REFACTORED EXPORT LOGIC ---
        
        # 1. Define the mapping from styled header name to data key
        header_map = {
            'First Name': 'FirstName', 'Last Name': 'LastName', 'Email': 'Email', 'Phone': 'PhoneNumber',
            'Gender': 'Gender', 'DOB': 'DateOfBirth', 'Nationality': 'CandidateNationality', 
            'Education': 'EducationalStatus', 'Languages': 'LanguagesStr', 'Lang. Level': 'LanguageLevel', 
            'LinkedIn': 'LinkedInProfileURL', 'Registered On': 'RegistrationDate', 'Company': 'CompanyName', 
            'Applying For': 'OfferTitle', 'Interview Date': 'InterviewDateStr', 'Interview Time': 'InterviewTimeStr'
        }
        
        # 2. Pre-process the data to match the keys in header_map
        excel_data = []
        for interview in scheduled_interviews:
            new_row = interview.copy()
            new_row['LanguagesStr'] = ", ".join(sorted(interview['Languages'])) if interview.get('Languages') else 'N/A'
            new_row['InterviewDateStr'] = interview['ScheduledDateTime'].strftime('%Y-%m-%d')
            new_row['InterviewTimeStr'] = interview['ScheduledDateTime'].strftime('%I:%M %p')
            excel_data.append(new_row)

        # 3. Call the helper function
        report_title = "Scheduled Interviews"
        if company_id_filter and company_id_filter.isdigit() and scheduled_interviews:
             report_title = f"Scheduled Interviews for {scheduled_interviews[0]['CompanyName']}"
        
        excel_file = _create_styled_excel(excel_data, report_title, header_map)

        # 4. Create the response
        filename = "scheduled_interviews.xlsx"
        if company_id_filter and company_id_filter.isdigit() and scheduled_interviews:
             company_name_slug = scheduled_interviews[0]['CompanyName'].replace(' ', '_').lower()
             filename = f"interviews_{company_name_slug}.xlsx"

        response = make_response(excel_file.read())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response

    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()