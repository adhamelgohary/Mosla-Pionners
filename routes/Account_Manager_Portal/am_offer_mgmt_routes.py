# routes/Account_Manager_Portal/am_offer_mgmt_routes.py
from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, make_response
from flask_login import current_user
from utils.decorators import login_required_with_role
from db import get_db_connection
import datetime
import decimal
import re
import io
import csv
from collections import OrderedDict

# Imports for styled Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Roles that can manage offers within the AM portal
AM_OFFER_MANAGEMENT_ROLES = ['HeadAccountManager', 'CEO', 'Founder', 'Admin']
CLIENT_CONTACT_ROLE_NAME = 'ClientContact' # Needed for client submission route if you add it here

am_offer_mgmt_bp = Blueprint('am_offer_mgmt_bp', __name__,
                              template_folder='../../../templates',
                              url_prefix='/am-portal/offer-management')

# --- Helper Functions (Copied from job_offer_mgmt_routes.py) ---
def _create_styled_excel(report_data, title, header_mapping):
    """Generates a styled Excel report using only OpenPyXL."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    title_font = Font(name='Calibri', size=18, bold=True, color='1F2937')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header_mapping))
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align
    worksheet.row_dimensions[1].height = 30
    
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header_mapping))
    subtitle_cell = worksheet.cell(row=2, column=1, value=f"Report generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    subtitle_cell.font = Font(italic=True, color='6B7280')
    subtitle_cell.alignment = center_align
    worksheet.row_dimensions[2].height = 20
    
    headers = list(header_mapping.keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    worksheet.row_dimensions[4].height = 20
    
    data_keys = list(header_mapping.values())
    for row_num, row_data in enumerate(report_data, 5):
        for col_num, key in enumerate(data_keys, 1):
            cell_value = row_data.get(key, 'N/A')
            if isinstance(cell_value, (datetime.datetime, datetime.date)):
                cell_value = cell_value.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=col_num, value=cell_value).alignment = left_align
    
    column_widths = {}
    for col_num, header_title in enumerate(headers, 1):
        column_widths[col_num] = len(header_title)
    for row_data in report_data:
        for col_num, key in enumerate(data_keys, 1):
            cell_len = len(str(row_data.get(key, '')))
            if cell_len > column_widths[col_num]:
                column_widths[col_num] = cell_len
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(col_num)].width = width + 4
        
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def get_column_options(conn, cursor, table_name, column_name):
    """Dynamically fetches the allowed values for an ENUM or SET column from the database schema."""
    try:
        db_name = conn.database 
        query = """
            SELECT COLUMN_TYPE 
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s AND COLUMN_NAME = %s
        """
        cursor.execute(query, (db_name, table_name, column_name))
        result = cursor.fetchone()
        if result and result.get('COLUMN_TYPE'):
            type_string = result['COLUMN_TYPE']
            values = re.findall(r"'(.*?)'", type_string)
            return values
        return []
    except Exception as e:
        current_app.logger.error(f"Could not fetch options for {table_name}.{column_name}: {e}", exc_info=True)
        return []

def validate_job_offer_data(form_data, is_editing=False):
    """A comprehensive validation helper for the offer form."""
    errors = {}
    if not form_data.get('title', '').strip():
        errors['title'] = 'Job title is required.'
    
    if not is_editing:
        company_selection_mode = form_data.get('company_selection_mode')
        if not form_data.get('company_id'):
            if company_selection_mode == 'existing' and not form_data.get('selected_company_id'):
                errors['company_id'] = "You must select an existing company."
            elif company_selection_mode == 'new' and not form_data.get('new_company_name', '').strip():
                errors['new_company_name'] = "New company name cannot be empty."
            elif not company_selection_mode:
                errors['company_selection_mode'] = "Please select a company option."
    
    if not form_data.get('category_id'):
        errors['category_id'] = 'Job category is required.'
    
    return errors
    
def _create_automated_announcement(db_cursor, source_type, title, content, audience='AllUsers', priority='Normal', posted_by_user_id=None):
    """Helper to create an automated announcement."""
    try:
        sql = "INSERT INTO SystemAnnouncements (Title, Content, Audience, Priority, Source, PostedByUserID, IsActive) VALUES (%s, %s, %s, %s, %s, %s, 1)"
        db_cursor.execute(sql, (title, content, audience, priority, source_type, posted_by_user_id))
        return True
    except Exception as e:
        current_app.logger.error(f"Failed to create automated announcement: {e}", exc_info=True)
        return False

# --- Main Routes ---

@am_offer_mgmt_bp.route('/')
@login_required_with_role(AM_OFFER_MANAGEMENT_ROLES)
def list_all_job_offers():
    """Main landing page for AM Offer Management, listing all offers grouped by company."""
    offers_by_company = OrderedDict()
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                jo.OfferID, jo.Title, jo.Status, jo.DatePosted, 
                c.CompanyID, c.CompanyName, 
                jc.CategoryName 
            FROM JobOffers jo 
            JOIN Companies c ON jo.CompanyID = c.CompanyID 
            LEFT JOIN JobCategories jc ON jo.CategoryID = jc.CategoryID 
            ORDER BY c.CompanyName, jo.DatePosted DESC
        """)
        all_offers = cursor.fetchall()

        for offer in all_offers:
            company_id = offer['CompanyID']
            if company_id not in offers_by_company:
                offers_by_company[company_id] = {
                    "company_id": company_id,
                    "company_name": offer['CompanyName'],
                    "offers": []
                }
            offers_by_company[company_id]['offers'].append(offer)
    except Exception as e:
        current_app.logger.error(f"Error fetching AM job offers: {e}", exc_info=True)
        flash("Could not load job offers list.", "danger")
    finally:
        if conn and conn.is_connected(): conn.close()
    
    return render_template('account_manager_portal/offers/list_all_job_offers.html', 
        title='Manage Live Offers', 
        offers_by_company=offers_by_company
    )


@am_offer_mgmt_bp.route('/create-live', methods=['GET', 'POST'])
@login_required_with_role(AM_OFFER_MANAGEMENT_ROLES)
def staff_direct_create_job_offer():
    errors, companies, categories, form_options = {}, [], [], {}
    form_data = {}

    conn_data = get_db_connection()
    try:
        cursor = conn_data.cursor(dictionary=True)
        cursor.execute("SELECT CompanyID, CompanyName FROM Companies ORDER BY CompanyName")
        companies = cursor.fetchall()
        cursor.execute("SELECT CategoryID, CategoryName FROM JobCategories ORDER BY CategoryName")
        categories = cursor.fetchall()
        
        # This function fetches ENUM/SET options
        def get_form_options(conn, cursor):
            options = {}
            table = 'JobOffers'
            columns_to_fetch = [
                'HiringPlan', 'GraduationStatusRequirement', 'HiringCadence',
                'WorkLocationType', 'ShiftType', 'AvailableShifts', 'BenefitsIncluded',
                'Nationality', 'RequiredLanguages', 'RequiredLevel', 'Status', 'PaymentTerm',
                'LanguagesType', 'Gender', 'MilitaryStatus'
            ]
            for col in columns_to_fetch:
                options[col] = get_column_options(conn, cursor, table, col)
            return options
        form_options = get_form_options(conn_data, cursor)

    finally:
        if conn_data and conn_data.is_connected(): conn_data.close()

    if request.method == 'POST':
        form_data = request.form.to_dict()
        form_data['benefits_included'] = request.form.getlist('benefits_checkboxes')
        form_data['available_shifts'] = request.form.getlist('available_shifts')
        form_data['required_languages'] = request.form.getlist('required_languages')
        form_data['grad_status_req'] = request.form.getlist('grad_status_req')
        errors = validate_job_offer_data(form_data, is_editing=False)

        if not errors:
            conn_tx = get_db_connection()
            try:
                cursor = conn_tx.cursor(dictionary=True)
                conn_tx.start_transaction()
                
                company_id, company_name = None, ''
                if form_data.get('company_selection_mode') == 'new':
                    new_company_name = form_data.get('new_company_name').strip()
                    cursor.execute("INSERT INTO Companies (CompanyName) VALUES (%s)", (new_company_name,))
                    company_id = cursor.lastrowid
                    company_name = new_company_name
                else:
                    company_id = form_data.get('selected_company_id')
                    cursor.execute("SELECT CompanyName FROM Companies WHERE CompanyID = %s", (company_id,))
                    company_name = cursor.fetchone()['CompanyName']

                params = {
                    "company_id": company_id,
                    "posted_by_id": getattr(current_user, 'specific_role_id', None),
                    "category_id": form_data.get('category_id'),
                    "title": form_data.get('title'), "location": form_data.get('location'),
                    "net_salary": decimal.Decimal(form_data['net_salary']) if form_data.get('net_salary') else None,
                    "payment_term": form_data.get('payment_term'), "hiring_plan": form_data.get('hiring_plan'),
                    "max_age": int(form_data['max_age']) if form_data.get('max_age') else None,
                    "has_contract": 1 if form_data.get('has_contract') == 'yes' else 0,
                    "grad_status_req": ",".join(form_data.get('grad_status_req', [])),
                    "languages_type": form_data.get('languages_type'), "required_languages": ",".join(form_data.get('required_languages', [])),
                    "required_level": form_data.get('required_level'), "candidates_needed": int(form_data.get('candidates_needed', 1)),
                    "hiring_cadence": form_data.get('hiring_cadence'), "work_location_type": form_data.get('work_location_type'),
                    "shift_type": form_data.get('shift_type'), "available_shifts": ",".join(form_data.get('available_shifts', [])),
                    "benefits_included": ",".join(form_data.get('benefits_included', [])),
                    "interview_type": form_data.get('interview_type'), "nationality": form_data.get('nationality'),
                    "closing_date": form_data.get('closing_date') or None, "gender": form_data.get('gender'),
                    "military_status": form_data.get('military_status'), "working_days": form_data.get('working_days'),
                    "working_hours": form_data.get('working_hours'), "experience_requirement": form_data.get('experience_requirement')
                }
                
                sql = """INSERT INTO JobOffers (
                        CompanyID, PostedByStaffID, CategoryID, Title, Location, NetSalary, PaymentTerm, HiringPlan,
                        MaxAge, HasContract, GraduationStatusRequirement, LanguagesType, RequiredLanguages, RequiredLevel,
                        CandidatesNeeded, HiringCadence, WorkLocationType, ShiftType, AvailableShifts,
                        BenefitsIncluded, InterviewType, Nationality, Status, ClosingDate, DatePosted,
                        Gender, MilitaryStatus, WorkingDays, WorkingHours, ExperienceRequirement
                    ) VALUES (
                        %(company_id)s, %(posted_by_id)s, %(category_id)s, %(title)s, %(location)s, %(net_salary)s, 
                        %(payment_term)s, %(hiring_plan)s, %(max_age)s, %(has_contract)s, %(grad_status_req)s,
                        %(languages_type)s, %(required_languages)s, %(required_level)s, %(candidates_needed)s, 
                        %(hiring_cadence)s, %(work_location_type)s, %(shift_type)s, %(available_shifts)s, 
                        %(benefits_included)s, %(interview_type)s, %(nationality)s, 'Open', %(closing_date)s, NOW(),
                        %(gender)s, %(military_status)s, %(working_days)s, %(working_hours)s, %(experience_requirement)s
                    )"""
                cursor.execute(sql, params)
                
                _create_automated_announcement(cursor, 'Automated_Offer', f"New Job: {params['title']} with {company_name}", "A new job offer has been posted.", posted_by_user_id=current_user.id)
                conn_tx.commit()
                flash(f"New job offer '{params['title']}' created and announced!", 'success')
                return redirect(url_for('.list_all_job_offers'))
            except Exception as e:
                if conn_tx: conn_tx.rollback()
                flash(f'An error occurred: {e}', 'danger')
            finally:
                if conn_tx and conn_tx.is_connected(): conn_tx.close()

    return render_template('account_manager_portal/offers/add_edit_job_offer.html', 
        title='Create Job Offer', form_data=form_data, errors=errors, companies=companies, 
        categories=categories, is_editing_live=False, action_verb="Post Live", form_options=form_options)


@am_offer_mgmt_bp.route('/edit-live/<int:offer_id>', methods=['GET', 'POST'])
@login_required_with_role(AM_OFFER_MANAGEMENT_ROLES)
def edit_live_job_offer(offer_id):
    # This function is a direct copy of the one from job_offer_mgmt_routes.py,
    # but redirects and renders templates appropriate for the AM portal.
    # The full logic is included for completeness.
    errors, companies, categories, form_options = {}, [], [], {}
    form_data_for_template = {}
    
    conn_deps = get_db_connection()
    try:
        cursor_deps = conn_deps.cursor(dictionary=True)
        cursor_deps.execute("SELECT CompanyID, CompanyName FROM Companies ORDER BY CompanyName")
        companies = cursor_deps.fetchall()
        cursor_deps.execute("SELECT CategoryID, CategoryName FROM JobCategories ORDER BY CategoryName")
        categories = cursor_deps.fetchall()
        def get_form_options(conn, cursor):
            options = {}
            table = 'JobOffers'
            columns_to_fetch = [
                'HiringPlan', 'GraduationStatusRequirement', 'HiringCadence',
                'WorkLocationType', 'ShiftType', 'AvailableShifts', 'BenefitsIncluded',
                'Nationality', 'RequiredLanguages', 'RequiredLevel', 'Status', 'PaymentTerm',
                'LanguagesType', 'Gender', 'MilitaryStatus'
            ]
            for col in columns_to_fetch:
                options[col] = get_column_options(conn, cursor, table, col)
            return options
        form_options = get_form_options(conn_deps, cursor_deps)
    finally:
        if conn_deps and conn_deps.is_connected(): conn_deps.close()

    conn_orig = get_db_connection()
    try:
        cursor_orig = conn_orig.cursor(dictionary=True)
        cursor_orig.execute("SELECT Title, Status FROM JobOffers WHERE OfferID = %s", (offer_id,))
        original_offer = cursor_orig.fetchone()
        if not original_offer:
            flash('Job offer not found.', 'danger'); return redirect(url_for('.list_all_job_offers'))
        old_status = original_offer['Status']
        original_title = original_offer['Title']
    finally:
        if conn_orig and conn_orig.is_connected(): conn_orig.close()

    if request.method == 'POST':
        form_data = request.form.to_dict()
        form_data['benefits_included'] = request.form.getlist('benefits_checkboxes')
        form_data['available_shifts'] = request.form.getlist('available_shifts')
        form_data['required_languages'] = request.form.getlist('required_languages')
        form_data['grad_status_req'] = request.form.getlist('grad_status_req')
        errors = validate_job_offer_data(form_data, is_editing=True)
        
        if not errors:
            conn_update = get_db_connection()
            try:
                cursor_update = conn_update.cursor()
                conn_update.start_transaction()
                
                params = {
                    "category_id": form_data.get('category_id'), "title": form_data.get('title').strip(), "location": form_data.get('location'),
                    "net_salary": decimal.Decimal(form_data['net_salary']) if form_data.get('net_salary') else None,
                    "payment_term": form_data.get('payment_term'), "hiring_plan": form_data.get('hiring_plan'),
                    "max_age": int(form_data['max_age']) if form_data.get('max_age') else None,
                    "has_contract": 1 if form_data.get('has_contract') == 'yes' else 0,
                    "grad_status_req": ",".join(form_data.get('grad_status_req', [])), "languages_type": form_data.get('languages_type'),
                    "required_languages": ",".join(form_data.get('required_languages', [])), "required_level": form_data.get('required_level'),
                    "candidates_needed": int(form_data.get('candidates_needed', 1)), "hiring_cadence": form_data.get('hiring_cadence'),
                    "work_location_type": form_data.get('work_location_type'), "shift_type": form_data.get('shift_type'),
                    "available_shifts": ",".join(form_data.get('available_shifts', [])), "benefits_included": ",".join(form_data.get('benefits_included', [])),
                    "interview_type": form_data.get('interview_type'), "nationality": form_data.get('nationality'),
                    "status": form_data.get('status', 'Open'), "closing_date": form_data.get('closing_date') or None,
                    "company_id": form_data.get('CompanyID'), "gender": form_data.get('gender'),
                    "military_status": form_data.get('military_status'), "working_days": form_data.get('working_days'),
                    "working_hours": form_data.get('working_hours'), "experience_requirement": form_data.get('experience_requirement'),
                    "offer_id": offer_id
                }
                
                sql = """UPDATE JobOffers SET
                        CompanyID=%(company_id)s, CategoryID=%(category_id)s, Title=%(title)s, Location=%(location)s, NetSalary=%(net_salary)s, PaymentTerm=%(payment_term)s, HiringPlan=%(hiring_plan)s,
                        MaxAge=%(max_age)s, HasContract=%(has_contract)s, GraduationStatusRequirement=%(grad_status_req)s, LanguagesType=%(languages_type)s, 
                        RequiredLanguages=%(required_languages)s, RequiredLevel=%(required_level)s, CandidatesNeeded=%(candidates_needed)s, 
                        HiringCadence=%(hiring_cadence)s, WorkLocationType=%(work_location_type)s, ShiftType=%(shift_type)s, 
                        AvailableShifts=%(available_shifts)s, BenefitsIncluded=%(benefits_included)s, InterviewType=%(interview_type)s, 
                        Nationality=%(nationality)s, Status=%(status)s, ClosingDate=%(closing_date)s, 
                        Gender=%(gender)s, MilitaryStatus=%(military_status)s, WorkingDays=%(working_days)s, WorkingHours=%(working_hours)s, ExperienceRequirement=%(experience_requirement)s,
                        UpdatedAt=NOW() WHERE OfferID = %(offer_id)s"""
                cursor_update.execute(sql, params)

                new_status = params['status']
                if old_status != new_status:
                    _create_automated_announcement(cursor_update, 'Automated_Status_Change', f"Offer Status Updated: {original_title}", f"The status for the job offer '{original_title}' has been changed from '{old_status}' to '{new_status}'.", posted_by_user_id=current_user.id)
                    flash(f"Job offer status updated and an announcement was posted.", 'info')

                conn_update.commit()
                flash(f"Job offer '{params['title']}' updated successfully!", 'success')
                return redirect(url_for('.list_all_job_offers'))
            except Exception as e:
                if conn_update: conn_update.rollback()
                flash(f'An error occurred: {e}', 'danger')
            finally:
                if conn_update and conn_update.is_connected(): conn_update.close()
        
        form_data_for_template = form_data
    
    else: # GET request
        conn_fetch = get_db_connection()
        try:
            cursor_fetch = conn_fetch.cursor(dictionary=True)
            cursor_fetch.execute("SELECT * FROM JobOffers WHERE OfferID = %s", (offer_id,))
            db_data = cursor_fetch.fetchone()
            form_data_for_template = db_data.copy()
            def parse_set_or_text_column(value):
                if isinstance(value, (bytes, str)): return [v.strip() for v in value.split(',') if v.strip()]
                return []
            form_data_for_template['required_languages'] = parse_set_or_text_column(db_data.get('RequiredLanguages'))
            form_data_for_template['available_shifts'] = parse_set_or_text_column(db_data.get('AvailableShifts'))
            form_data_for_template['grad_status_req'] = parse_set_or_text_column(db_data.get('GraduationStatusRequirement'))
            form_data_for_template['benefits_included'] = parse_set_or_text_column(db_data.get('BenefitsIncluded'))
            form_data_for_template['has_contract'] = 'yes' if db_data.get('HasContract') else 'no'
            form_data_for_template['net_salary'] = str(db_data['NetSalary']) if db_data.get('NetSalary') is not None else ''
            form_data_for_template['closing_date'] = db_data['ClosingDate'].isoformat() if db_data.get('ClosingDate') else ''
        finally: 
            if conn_fetch and conn_fetch.is_connected(): conn_fetch.close()
                
    return render_template('account_manager_portal/offers/add_edit_job_offer.html', 
        title=f'Edit Offer: {form_data_for_template.get("Title", "N/A")}', 
        form_data=form_data_for_template, errors=errors, companies=companies, categories=categories, 
        offer_id=offer_id, is_editing_live=True, action_verb="Update Offer", form_options=form_options)


@am_offer_mgmt_bp.route('/delete-live/<int:offer_id>', methods=['POST'])
@login_required_with_role(AM_OFFER_MANAGEMENT_ROLES)
def delete_live_job_offer(offer_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM JobOffers WHERE OfferID = %s", (offer_id,))
        conn.commit()
        if cursor.rowcount > 0:
            flash('The job offer has been permanently deleted.', 'success')
        else:
            flash('The job offer could not be found.', 'warning')
    except Exception as e:
        if conn: conn.rollback()
        flash('Cannot delete this offer because it has active job applications.', 'danger')
    finally:
        if conn and conn.is_connected(): conn.close()
    return redirect(url_for('.list_all_job_offers'))